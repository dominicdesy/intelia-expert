#!/usr/bin/env python3
"""
PDF Table Extractor to Excel
=============================

Extrait tous les tableaux d'un PDF Cobb et les transforme en fichier Excel structuré
avec métadonnées complètes.

Usage:
    python extract_pdf_tables_to_excel.py <pdf_file> <output_excel>

Example:
    python extract_pdf_tables_to_excel.py Cobb500-2022.pdf output.xlsx
"""

import pdfplumber
import pandas as pd
import sys
import re
from pathlib import Path
from typing import List, Dict, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


class CobbTableExtractor:
    """Extracteur de tableaux pour documents Cobb"""

    def __init__(self, pdf_path: str):
        self.pdf_path = Path(pdf_path)
        self.pdf = None
        self.tables = []

    def open_pdf(self):
        """Ouvre le PDF"""
        self.pdf = pdfplumber.open(self.pdf_path)
        print(f"[OK] PDF ouvert: {self.pdf_path.name}")
        print(f"  Pages: {len(self.pdf.pages)}")

    def close_pdf(self):
        """Ferme le PDF"""
        if self.pdf:
            self.pdf.close()

    def merge_table_fragments(self, fragments: List[List]) -> List[List]:
        """
        Fusionne les fragments de tableau en un seul tableau complet.
        PDFPlumber fragmente souvent les tableaux ligne par ligne.
        """
        if not fragments:
            return []

        # Le premier fragment contient généralement l'en-tête
        merged = []
        for fragment in fragments:
            if fragment:  # Ignorer les fragments vides
                merged.extend(fragment)

        return merged

    def _detect_header_rows(self, table: List[List]) -> Tuple[int, List[str]]:
        """
        Détecte intelligemment les lignes d'en-tête et retourne l'index et les noms fusionnés

        Returns:
            (header_end_index, clean_header_names)
        """
        # Chercher la première ligne qui contient des mots-clés typiques d'en-tête
        header_keywords = ['age', 'weight', 'gain', 'feed', 'nutrient', 'amino', 'yield', 'days', 'protein', 'energy']

        header_start = -1
        for i, row in enumerate(table):
            if not row:
                continue
            row_text = ' '.join([str(cell).lower() for cell in row if cell])
            if any(keyword in row_text for keyword in header_keywords):
                header_start = i
                break

        if header_start == -1:
            return -1, []

        # Vérifier si l'en-tête est multi-ligne (ligne suivante contient des unités comme g, kg, %, days)
        unit_keywords = ['(g)', '(kg)', '(lb)', '(%)', '(days)', 'g)', 'kg)', '%)', 'days)']
        has_unit_row = False
        header_end = header_start

        if header_start + 1 < len(table):
            next_row = table[header_start + 1]
            next_row_text = ' '.join([str(cell).lower() for cell in next_row if cell])
            if any(unit in next_row_text for unit in unit_keywords):
                has_unit_row = True
                header_end = header_start + 1

        # Fusionner les lignes d'en-tête
        clean_header = []
        header_row1 = table[header_start]

        if has_unit_row:
            header_row2 = table[header_start + 1]
            for name, unit in zip(header_row1, header_row2):
                name_str = str(name).replace('\n', ' ').strip() if name else ''
                unit_str = str(unit).replace('\n', ' ').strip() if unit else ''

                if name_str and unit_str:
                    clean_header.append(f"{name_str} {unit_str}")
                elif name_str:
                    clean_header.append(name_str)
                elif unit_str:
                    clean_header.append(unit_str)
                else:
                    clean_header.append('')
        else:
            clean_header = [str(col).replace('\n', ' ').strip() if col else '' for col in header_row1]

        return header_end, clean_header

    def _detect_data_rows(self, table: List[List], start_index: int) -> List[List]:
        """
        Extrait les lignes de données à partir d'un index de départ
        S'arrête aux lignes vides ou aux notes de bas de page
        """
        data_rows = []

        for i in range(start_index + 1, len(table)):
            row = table[i]

            # Ignorer les lignes totalement vides
            if not row or all(not cell or str(cell).strip() == '' for cell in row):
                continue

            # Vérifier si la première cellule contient un nombre (indique une ligne de données)
            first_cell = str(row[0]).strip() if row[0] else ''

            # Si la première cellule est vide, on continue (peut être une ligne vide intermédiaire)
            if not first_cell:
                continue

            # Si la première cellule est un nombre, c'est probablement une ligne de données
            if first_cell.replace('.', '').replace(',', '').isdigit():
                data_rows.append(row)
            else:
                # Si on a déjà des données et qu'on rencontre du texte, c'est probablement une note
                if len(data_rows) > 0:
                    break

        return data_rows

    def extract_performance_table(self, page_num: int) -> Dict:
        """
        Extrait une table de performance avec détection automatique de la structure
        """
        page = self.pdf.pages[page_num]

        # Use extract_table (singular) with text strategy to capture all rows
        table = page.extract_table(table_settings={
            "vertical_strategy": "text",
            "horizontal_strategy": "text",
        })

        if not table or len(table) < 3:
            return None

        # Détecter l'en-tête automatiquement
        header_end_idx, clean_header = self._detect_header_rows(table)

        if header_end_idx == -1:
            return None

        # Extraire les données automatiquement
        data_rows = self._detect_data_rows(table, header_end_idx)

        if not data_rows:
            return None

        df = pd.DataFrame(data_rows, columns=clean_header)

        # Nettoyer les données
        df = self._clean_dataframe(df)

        return {
            'type': 'performance',
            'page': page_num + 1,
            'header': clean_header,
            'data': df,
            'raw_rows': len(data_rows)
        }

    def extract_nutrition_table(self, page_num: int) -> Dict:
        """
        Extrait une table de nutrition (pages 10-12 typiquement)

        Structure:
        - Nutrient name | Unit | Starter | Grower | Finisher | Withdrawal
        """
        page = self.pdf.pages[page_num]

        # Use extract_table with text strategy
        table = page.extract_table(table_settings={
            "vertical_strategy": "text",
            "horizontal_strategy": "text",
        })

        if not table or len(table) < 2:
            return None

        # Nettoyer l'en-tête
        header = [str(col).replace('\n', ' ').strip() if col else '' for col in table[0]]
        data_rows = table[1:]

        df = pd.DataFrame(data_rows, columns=header)
        df = self._clean_dataframe(df)

        return {
            'type': 'nutrition',
            'page': page_num + 1,
            'header': header,
            'data': df,
            'raw_rows': len(table)
        }

    def extract_amino_acid_table(self, page_num: int, table_index: int = 1) -> Dict:
        """
        Extrait une table d'acides aminés (souvent 2ème table sur pages 10-11)

        Structure:
        - Amino Acid | Starter % | Grower % | Finisher % | Withdrawal %
        """
        page = self.pdf.pages[page_num]

        # For pages with multiple tables, use extract_tables() but with text strategy
        tables = page.extract_tables(table_settings={
            "vertical_strategy": "text",
            "horizontal_strategy": "text",
        })

        if len(tables) < table_index + 1:
            return None

        amino_table = tables[table_index]

        if len(amino_table) < 2:
            return None

        header = [str(col).replace('\n', ' ').strip() if col else '' for col in amino_table[0]]
        data_rows = amino_table[1:]

        df = pd.DataFrame(data_rows, columns=header)
        df = self._clean_dataframe(df)

        return {
            'type': 'amino_acids',
            'page': page_num + 1,
            'header': header,
            'data': df,
            'raw_rows': len(amino_table)
        }

    def extract_yield_table(self, page_num: int) -> Dict:
        """
        Extrait une table de rendement (pages 13-14)

        Structure:
        - Live Weight | Carcass | Breast | Legs | Wings | etc.
        """
        page = self.pdf.pages[page_num]

        # Use extract_table with text strategy
        table = page.extract_table(table_settings={
            "vertical_strategy": "text",
            "horizontal_strategy": "text",
        })

        if not table or len(table) < 2:
            return None

        header = [str(col).replace('\n', ' ').strip() if col else '' for col in table[0]]
        data_rows = table[1:]

        df = pd.DataFrame(data_rows, columns=header)
        df = self._clean_dataframe(df)

        return {
            'type': 'yield',
            'page': page_num + 1,
            'header': header,
            'data': df,
            'raw_rows': len(table)
        }

    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Nettoie un DataFrame"""
        # Supprimer les lignes totalement vides
        df = df.dropna(how='all')

        # Supprimer les colonnes totalement vides
        df = df.dropna(axis=1, how='all')

        # Supprimer les colonnes sans nom
        df = df.loc[:, df.columns != '']

        # Nettoyer les valeurs
        df = df.map(lambda x: str(x).strip() if x and str(x).strip() != 'None' else '')

        # Reset index
        df = df.reset_index(drop=True)

        return df

    def detect_table_metadata(self, df: pd.DataFrame, table_type: str) -> Dict:
        """
        Détecte les métadonnées d'une table basé sur son contenu
        """
        metadata = {
            'brand': 'cobb-vantress',
            'breed': 'cobb',
            'strain': '500',
            'type': 'commercial broiler chicken',
            'bird_type': 'broiler',
            'year': '2022'
        }

        # Détecter le sexe basé sur les données
        if 'Weight' in df.columns and not df.empty:
            # Analyser le poids à 56 jours pour deviner le sexe
            try:
                weight_col = [col for col in df.columns if 'Weight' in col or 'weight' in col][0]
                last_weight = float(df[df.iloc[:, 0] == '56'][weight_col].iloc[0]) if not df[df.iloc[:, 0] == '56'].empty else 0

                if last_weight > 4900:
                    metadata['sex'] = 'male'
                elif last_weight < 4400:
                    metadata['sex'] = 'female'
                else:
                    metadata['sex'] = 'mixed'
            except:
                metadata['sex'] = 'mixed'
        else:
            metadata['sex'] = 'mixed'

        # Description basée sur le type
        if table_type == 'performance':
            metadata['description'] = f"performance objectives for {metadata['sex']} broiler chickens in metric units"
            metadata['life_stage'] = 'growth period 0-56 days'
            metadata['age_days'] = '0-56'
        elif table_type == 'nutrition':
            metadata['description'] = "recommended nutrient levels for medium and large broiler chickens"
            metadata['life_stage'] = 'complete production cycle'
            metadata['age_days'] = '0-end of production'
        elif table_type == 'amino_acids':
            metadata['description'] = "balanced digestible amino acid ratios for medium and large broiler chickens"
            metadata['data_type'] = 'amino acid ratio specifications'
            metadata['life_stage'] = 'complete production cycle'
            metadata['age_days'] = '0-end of production'
        elif table_type == 'yield':
            metadata['description'] = f"carcass and cut-up yield percentages for {metadata['sex']} broiler chickens"
            metadata['life_stage'] = 'processing weights'
            metadata['age_days'] = 'variable processing age'

        return metadata

    def extract_all_tables(self) -> List[Dict]:
        """Extrait tous les tableaux du PDF"""
        self.open_pdf()

        all_tables = []

        try:
            # Pages 4-6: Performance tables (mixed, male, female)
            performance_pages = [3, 4, 5]  # Index 0-based
            for i, page_idx in enumerate(performance_pages):
                print(f"\n{'='*60}")
                print(f"Extraction page {page_idx + 1} - Performance table {i+1}/3...")
                table = self.extract_performance_table(page_idx)
                if table:
                    all_tables.append(table)
                    print(f"  [OK] Extrait: {len(table['data'])} lignes de données")

            # Pages 7-9: Performance tables en unités impériales (on peut les ignorer si metric seulement)
            # Pages 10-11: Nutrition + Amino acids
            print(f"\n{'='*60}")
            print(f"Extraction page 10 - Nutrition table...")
            nutr_table = self.extract_nutrition_table(9)
            if nutr_table:
                all_tables.append(nutr_table)
                print(f"  [OK] Extrait: {len(nutr_table['data'])} lignes")

            print(f"\nExtraction page 10 - Amino acids table...")
            amino_table = self.extract_amino_acid_table(9, table_index=1)
            if amino_table:
                all_tables.append(amino_table)
                print(f"  [OK] Extrait: {len(amino_table['data'])} lignes")

            # Page 12: Small chicken nutrition
            print(f"\n{'='*60}")
            print(f"Extraction page 12 - Small chicken nutrition...")
            small_nutr = self.extract_nutrition_table(11)
            if small_nutr:
                small_nutr['type'] = 'nutrition_small'
                all_tables.append(small_nutr)
                print(f"  [OK] Extrait: {len(small_nutr['data'])} lignes")

            # Pages 13-14: Yield tables
            print(f"\n{'='*60}")
            yield_pages = [12, 13]
            for i, page_idx in enumerate(yield_pages):
                print(f"Extraction page {page_idx + 1} - Yield table {i+1}...")
                yield_table = self.extract_yield_table(page_idx)
                if yield_table:
                    all_tables.append(yield_table)
                    print(f"  [OK] Extrait: {len(yield_table['data'])} lignes")

        finally:
            self.close_pdf()

        print(f"\n{'='*60}")
        print(f"[OK] Total tableaux extraits: {len(all_tables)}")
        return all_tables

    def create_excel_sheet(self, workbook: openpyxl.Workbook, sheet_name: str,
                          metadata: Dict, df: pd.DataFrame, column_info: List[Dict]):
        """
        Crée une feuille Excel avec métadonnées et données

        Structure:
        - Lignes 1-40: Métadonnées (metadata, value)
        - Lignes 41+: Données de la table
        """
        ws = workbook.create_sheet(sheet_name)

        # Style pour les métadonnées
        metadata_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        metadata_header_font = Font(color="FFFFFF", bold=True)

        # Écrire les métadonnées
        row = 1
        ws.cell(row, 1, 'metadata').font = metadata_header_font
        ws.cell(row, 1).fill = metadata_header_fill
        ws.cell(row, 2, 'value').font = metadata_header_font
        ws.cell(row, 2).fill = metadata_header_fill
        row += 1

        for key, value in metadata.items():
            ws.cell(row, 1, key)
            ws.cell(row, 2, value)
            row += 1

        # Ajouter les informations de table
        ws.cell(row, 1, 'table_header_row'); ws.cell(row, 2, row + 3); row += 1
        ws.cell(row, 1, 'table_data_rows'); ws.cell(row, 2, len(df)); row += 1
        ws.cell(row, 1, 'table_columns'); ws.cell(row, 2, len(df.columns)); row += 1
        ws.cell(row, 1, 'expected_metrics'); ws.cell(row, 2, len(df) * len(df.columns)); row += 1

        # Checksum
        ws.cell(row, 1, 'validation_checksum')
        ws.cell(row, 2, f"cobb500_{sheet_name}")
        row += 1

        # Informations sur les colonnes
        for i, col_info in enumerate(column_info, 1):
            ws.cell(row, 1, f'column_{i}_name'); ws.cell(row, 2, col_info['name']); row += 1
            ws.cell(row, 1, f'column_{i}_type'); ws.cell(row, 2, col_info['type']); row += 1
            ws.cell(row, 1, f'column_{i}_unit'); ws.cell(row, 2, col_info['unit']); row += 1

        # Ligne vide
        row += 2

        # Style pour l'en-tête de données
        data_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        data_header_font = Font(bold=True)

        # Écrire l'en-tête des données
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row, col_idx, col_name)
            cell.fill = data_header_fill
            cell.font = data_header_font
            cell.alignment = Alignment(horizontal='center')
        row += 1

        # Écrire les données
        for _, data_row in df.iterrows():
            for col_idx, value in enumerate(data_row, 1):
                ws.cell(row, col_idx, value)
            row += 1

        # Ajuster la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    def save_to_excel(self, tables: List[Dict], output_path: str):
        """Sauvegarde tous les tableaux dans un fichier Excel"""
        print(f"\n{'='*60}")
        print(f"Création du fichier Excel: {output_path}")

        wb = openpyxl.Workbook()
        # Supprimer la feuille par défaut
        wb.remove(wb.active)

        # Mappage des types de tables aux noms de feuilles
        sheet_names = {
            'performance': ['mixed_metric', 'male_metric', 'female_metric'],
            'nutrition': ['nutrient_med_large'],
            'nutrition_small': ['nutrient_small'],
            'amino_acids': ['amino_med_large', 'amino_small'],
            'yield': ['yield_mixed_metric', 'yield_male_metric', 'yield_female_metric']
        }

        sheet_index = {key: 0 for key in sheet_names}

        for table in tables:
            table_type = table['type']
            df = table['data']

            # Déterminer le nom de la feuille
            if table_type in sheet_names:
                names_list = sheet_names[table_type]
                idx = sheet_index[table_type]
                if idx < len(names_list):
                    sheet_name = names_list[idx]
                    sheet_index[table_type] += 1
                else:
                    sheet_name = f"{table_type}_{idx+1}"
            else:
                sheet_name = table_type

            # Détecter les métadonnées
            metadata = self.detect_table_metadata(df, table_type)

            # Inférer les types de colonnes
            column_info = []
            for col in df.columns:
                col_name = str(col).replace(' ', '_').replace('(', '').replace(')', '').replace('.', '')

                # Déterminer le type et l'unité
                if 'age' in col.lower() or 'day' in col.lower():
                    col_type, col_unit = 'integer', 'days'
                elif 'weight' in col.lower():
                    col_type, col_unit = 'integer', 'grams'
                elif 'gain' in col.lower():
                    col_type, col_unit = 'integer', 'grams'
                elif 'conversion' in col.lower() or 'fcr' in col.lower():
                    col_type, col_unit = 'numeric', 'ratio'
                elif 'intake' in col.lower():
                    col_type, col_unit = 'integer', 'grams'
                elif '%' in col:
                    col_type, col_unit = 'numeric', 'percent'
                else:
                    col_type, col_unit = 'text', 'various'

                column_info.append({
                    'name': col_name,
                    'type': col_type,
                    'unit': col_unit
                })

            # Créer la feuille
            print(f"  -> Creation feuille: {sheet_name}")
            self.create_excel_sheet(wb, sheet_name, metadata, df, column_info)

        # Sauvegarder
        wb.save(output_path)
        print(f"\n[OK] Fichier Excel sauvegarde: {output_path}")
        print(f"  Feuilles creees: {len(wb.sheetnames)}")


def main():
    """Fonction principale"""
    if len(sys.argv) < 3:
        print("Usage: python extract_pdf_tables_to_excel.py <pdf_file> <output_excel>")
        print("\nExample:")
        print("  python extract_pdf_tables_to_excel.py Cobb500-2022.pdf output.xlsx")
        sys.exit(1)

    pdf_file = sys.argv[1]
    output_excel = sys.argv[2]

    if not Path(pdf_file).exists():
        print(f"[ERROR] Le fichier {pdf_file} n'existe pas")
        sys.exit(1)

    print("="*60)
    print("PDF Table Extractor to Excel")
    print("="*60)

    extractor = CobbTableExtractor(pdf_file)

    try:
        # Extraire tous les tableaux
        tables = extractor.extract_all_tables()

        if not tables:
            print("[ERROR] Aucun tableau trouve dans le PDF")
            sys.exit(1)

        # Sauvegarder en Excel
        extractor.save_to_excel(tables, output_excel)

        print("\n" + "="*60)
        print("[OK] Extraction terminee avec succes!")
        print("="*60)

    except Exception as e:
        print(f"\n[ERROR] {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()

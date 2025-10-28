#!/usr/bin/env python3
"""
Merge multiple Excel files into one
"""

import openpyxl
import sys
from pathlib import Path


def merge_excel_files(source_files, output_file):
    """Fusionne plusieurs fichiers Excel en un seul"""

    # Créer un nouveau workbook
    output_wb = openpyxl.Workbook()
    output_wb.remove(output_wb.active)  # Supprimer la feuille par défaut

    sheet_names_seen = set()

    for source_file in source_files:
        if not Path(source_file).exists():
            print(f"[WARNING] Fichier {source_file} introuvable, ignoré")
            continue

        print(f"[Processing] {source_file}...")
        wb = openpyxl.load_workbook(source_file)

        for sheet_name in wb.sheetnames:
            source_sheet = wb[sheet_name]

            # Éviter les doublons de noms de feuilles
            final_sheet_name = sheet_name
            counter = 2
            while final_sheet_name in sheet_names_seen:
                final_sheet_name = f"{sheet_name}_{counter}"
                counter += 1

            sheet_names_seen.add(final_sheet_name)

            # Créer la nouvelle feuille
            new_sheet = output_wb.create_sheet(final_sheet_name)

            # Copier toutes les cellules
            for row in source_sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet[cell.coordinate]
                    new_cell.value = cell.value

                    # Copier le style
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.alignment = cell.alignment.copy()

            # Copier les largeurs de colonnes
            for col_letter, col_dim in source_sheet.column_dimensions.items():
                new_sheet.column_dimensions[col_letter].width = col_dim.width

            print(f"  -> Copié: {final_sheet_name}")

        wb.close()

    # Sauvegarder
    output_wb.save(output_file)
    print(f"\n[OK] Fichier fusionné: {output_file}")
    print(f"  Total feuilles: {len(output_wb.sheetnames)}")


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python merge_excel_files.py <output_file> <input_file1> <input_file2> ...")
        sys.exit(1)

    output_file = sys.argv[1]
    source_files = sys.argv[2:]

    merge_excel_files(source_files, output_file)

#!/usr/bin/env python3
"""
Module d'export simple et propre - CSV et XLSX
MODIFIÉ pour supporter ProcessedTableData
"""

import json
import csv
import re
from pathlib import Path
from typing import Dict, List, Any, Union
from dataclasses import asdict
import logging

# Import des types de données supportés
try:
    from parser import TableData, ProcessedTableData
except ImportError:
    # Fallback si les types ne sont pas disponibles
    TableData = None
    ProcessedTableData = None


class SimpleTableExporter:
    """Exporteur simple et propre - CSV et XLSX"""

    def __init__(self, output_dir: str):
        # Accepter n'importe quel chemin (relatif ou absolu)
        self.output_dir = Path(output_dir).resolve()
        self.logger = logging.getLogger(__name__)

        # Créer la structure de répertoires
        self._setup_output_structure()

        # Vérifier la disponibilité d'openpyxl
        self.xlsx_available = self._check_xlsx_support()

        # Ajouter warning si openpyxl non disponible
        if not self.xlsx_available:
            self.logger.warning(
                "openpyxl not available. Install with: pip install openpyxl"
            )

    def _setup_output_structure(self):
        """Créer la structure de répertoires de sortie"""
        try:
            self.output_dir.mkdir(parents=True, exist_ok=True)
            (self.output_dir / "csv").mkdir(exist_ok=True)
            (self.output_dir / "xlsx").mkdir(exist_ok=True)
            (self.output_dir / "metadata").mkdir(exist_ok=True)

            self.logger.info(f"Output directory: {self.output_dir}")

        except Exception as e:
            self.logger.error(f"Cannot create output directory {self.output_dir}: {e}")
            raise

    def _check_xlsx_support(self) -> bool:
        """Vérifie si openpyxl est disponible"""
        try:
            import importlib.util
            return importlib.util.find_spec("openpyxl") is not None
        except ImportError:
            return False

    def export_table(self, table: Union[TableData, ProcessedTableData, Any]) -> Dict[str, str]:
        """Exporte un tableau vers CSV et XLSX avec métadonnées séparées
        
        MODIFIÉ: Supporte maintenant TableData ET ProcessedTableData
        """
        
        # Normaliser les données selon le type
        normalized_table = self._normalize_table_data(table)
        
        # Générer nom de fichier basé sur les métadonnées
        filename_base = self._generate_filename(normalized_table)

        # Chemins de sortie
        csv_path = self.output_dir / "csv" / f"{filename_base}.csv"
        xlsx_path = self.output_dir / "xlsx" / f"{filename_base}.xlsx"
        metadata_path = self.output_dir / "metadata" / f"{filename_base}.json"

        results = {}

        # Export CSV propre
        try:
            self._export_csv(normalized_table, csv_path)
            results["csv"] = str(csv_path)
            self.logger.debug(f"CSV exported: {csv_path.name}")

        except Exception as e:
            self.logger.error(f"Error exporting CSV: {e}")
            raise

        # Export XLSX si disponible
        if self.xlsx_available:
            try:
                self._export_xlsx(normalized_table, xlsx_path)
                results["xlsx"] = str(xlsx_path)
                self.logger.debug(f"XLSX exported: {xlsx_path.name}")

            except Exception as e:
                self.logger.warning(f"Error exporting XLSX: {e}")
        else:
            self.logger.info("XLSX export skipped (openpyxl not available)")

        # Export métadonnées séparées
        try:
            self._export_metadata(normalized_table, metadata_path)
            results["metadata"] = str(metadata_path)
            self.logger.debug(f"Metadata exported: {metadata_path.name}")

        except Exception as e:
            self.logger.error(f"Error exporting metadata: {e}")
            raise

        return results

    def _normalize_table_data(self, table: Any) -> Dict[str, Any]:
        """Normalise les données de table selon le type d'entrée
        
        NOUVEAU: Gère TableData, ProcessedTableData et structures dict
        """
        
        # Détecter le type et extraire les données
        if hasattr(table, 'processed_records'):
            # Type ProcessedTableData
            return {
                'title': getattr(table, 'title', 'Unknown Table'),
                'headers': getattr(table, 'headers', []),
                'rows': getattr(table, 'raw_rows', []),  # Utiliser raw_rows
                'processed_records': getattr(table, 'processed_records', []),
                'metadata': getattr(table, 'metadata', {}),
                'is_pivot': getattr(table, 'is_pivot', False)
            }
        
        elif hasattr(table, 'rows'):
            # Type TableData classique
            return {
                'title': getattr(table, 'title', 'Unknown Table'),
                'headers': getattr(table, 'headers', []),
                'rows': getattr(table, 'rows', []),
                'processed_records': [],
                'metadata': getattr(table, 'metadata', {}),
                'is_pivot': False
            }
        
        elif isinstance(table, dict):
            # Structure dict directe
            return {
                'title': table.get('title', 'Unknown Table'),
                'headers': table.get('headers', []),
                'rows': table.get('rows', table.get('raw_rows', [])),
                'processed_records': table.get('processed_records', []),
                'metadata': table.get('metadata', {}),
                'is_pivot': table.get('is_pivot', False)
            }
        
        else:
            # Fallback - structure minimale
            self.logger.warning(f"Unknown table type: {type(table)}")
            return {
                'title': str(table) if hasattr(table, '__str__') else 'Unknown Table',
                'headers': [],
                'rows': [],
                'processed_records': [],
                'metadata': {},
                'is_pivot': False
            }

    def _export_csv(self, table_data: Dict[str, Any], csv_path: Path):
        """Export CSV - MODIFIÉ pour gérer les données normalisées"""
        
        headers = table_data['headers']
        rows = table_data['rows']
        processed_records = table_data['processed_records']
        
        # Si on a des enregistrements traités, les utiliser de préférence
        if processed_records and len(processed_records) > 0:
            self._export_processed_records_csv(processed_records, csv_path)
        else:
            # Export classique
            with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                if headers:
                    writer.writerow(headers)
                if rows:
                    writer.writerows(rows)

    def _export_processed_records_csv(self, records: List[Dict[str, Any]], csv_path: Path):
        """Export CSV pour les enregistrements traités (données dénormalisées)"""
        
        if not records:
            return
        
        # Obtenir toutes les clés possibles
        all_keys = set()
        for record in records:
            all_keys.update(record.keys())
        
        # Trier les headers pour un ordre logique
        sorted_headers = self._sort_headers_logically(list(all_keys))
        
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=sorted_headers)
            writer.writeheader()
            
            for record in records:
                # Compléter les valeurs manquantes
                complete_record = {key: record.get(key, "") for key in sorted_headers}
                writer.writerow(complete_record)

    def _sort_headers_logically(self, headers: List[str]) -> List[str]:
        """Trie les headers de manière logique pour faciliter la lecture"""
        
        # Ordre de priorité
        priority_prefixes = [
            'genetic_line', 'sex', 'bird_type', 'phase_name',
            'age_', 'amino_acid', 'nutrient', 'value_',
            'weight_', 'fcr_', 'feed_', 'gain_'
        ]
        
        priority_headers = []
        remaining_headers = []
        
        for header in headers:
            header_lower = header.lower()
            is_priority = False
            
            for prefix in priority_prefixes:
                if header_lower.startswith(prefix):
                    priority_headers.append(header)
                    is_priority = True
                    break
            
            if not is_priority:
                remaining_headers.append(header)
        
        # Trier chaque groupe
        priority_headers.sort()
        remaining_headers.sort()
        
        return priority_headers + remaining_headers

    def _export_xlsx(self, table_data: Dict[str, Any], xlsx_path: Path):
        """Export XLSX formaté avec styles - MODIFIÉ"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        # Créer le workbook
        workbook = openpyxl.Workbook()
        
        # Feuille principale avec données brutes
        worksheet = workbook.active
        worksheet.title = "Raw_Data"
        
        headers = table_data['headers']
        rows = table_data['rows']
        
        if headers and rows:
            self._write_xlsx_sheet(worksheet, headers, rows)
        
        # Feuille avec données traitées si disponibles
        processed_records = table_data['processed_records']
        if processed_records:
            processed_sheet = workbook.create_sheet("Processed_Data")
            self._write_processed_xlsx_sheet(processed_sheet, processed_records)
        
        # Feuille métadonnées
        self._add_metadata_sheet(workbook, table_data)
        
        # Sauvegarder
        workbook.save(xlsx_path)

    def _write_xlsx_sheet(self, worksheet, headers: List[str], rows: List[List[str]]):
        """Écrit une feuille XLSX avec formatage standard"""
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Écrire les headers avec formatage
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment

        # Écrire les données
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                if col_idx <= len(headers):  # Éviter les débordements
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    cell.border = border
                    
                    if self._is_numeric(cell_value):
                        cell.alignment = Alignment(horizontal="right")

        # Auto-ajustement des largeurs
        self._adjust_column_widths(worksheet)

    def _write_processed_xlsx_sheet(self, worksheet, records: List[Dict[str, Any]]):
        """Écrit une feuille XLSX avec les enregistrements traités"""
        if not records:
            return
        
        # Obtenir headers triés
        all_keys = set()
        for record in records:
            all_keys.update(record.keys())
        sorted_headers = self._sort_headers_logically(list(all_keys))
        
        # Écrire comme tableau standard
        rows_data = []
        for record in records:
            row = [record.get(key, "") for key in sorted_headers]
            rows_data.append(row)
        
        self._write_xlsx_sheet(worksheet, sorted_headers, rows_data)

    def _adjust_column_widths(self, worksheet):
        """Ajuste automatiquement les largeurs de colonnes"""
        from openpyxl.utils import get_column_letter
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, TypeError):
                    pass

            # Largeur ajustée avec limites
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Figer la ligne d'en-tête
        if worksheet.max_row > 1:
            worksheet.freeze_panes = worksheet["A2"]

    def _add_metadata_sheet(self, workbook, table_data: Dict[str, Any]):
        """Ajoute une feuille avec les métadonnées détaillées - MODIFIÉ"""
        try:
            from openpyxl.styles import Font
        except ImportError:
            return

        metadata_sheet = workbook.create_sheet("Metadata")
        title_font = Font(bold=True, size=14, color="366092")
        section_font = Font(bold=True, size=11)

        row = 1
        metadata = table_data.get('metadata', {})

        # Informations du tableau
        metadata_sheet.cell(row=row, column=1, value="INFORMATIONS DU TABLEAU").font = title_font
        row += 2

        basic_info = [
            ("Titre:", table_data.get('title', 'Unknown')),
            ("Nombre de colonnes:", len(table_data.get('headers', []))),
            ("Nombre de lignes:", len(table_data.get('rows', []))),
            ("Est une table pivot:", table_data.get('is_pivot', False)),
            ("Enregistrements traités:", len(table_data.get('processed_records', [])))
        ]

        for label, value in basic_info:
            metadata_sheet.cell(row=row, column=1, value=label).font = section_font
            metadata_sheet.cell(row=row, column=2, value=str(value))
            row += 1

        # Métadonnées normalisées si disponibles
        if metadata:
            row += 1
            metadata_sheet.cell(row=row, column=1, value="MÉTADONNÉES NORMALISÉES").font = title_font
            row += 2

            # Convertir métadonnées en dict si nécessaire
            if hasattr(metadata, '__dict__'):
                metadata_dict = metadata.__dict__
            elif hasattr(metadata, '_asdict'):
                metadata_dict = metadata._asdict()
            else:
                metadata_dict = metadata if isinstance(metadata, dict) else {}

            for key, value in metadata_dict.items():
                metadata_sheet.cell(row=row, column=1, value=f"{key}:").font = section_font
                metadata_sheet.cell(row=row, column=2, value=str(value))
                row += 1

        # Auto-ajustement métadonnées
        self._adjust_column_widths(metadata_sheet)

    def _export_metadata(self, table_data: Dict[str, Any], metadata_path: Path):
        """Export métadonnées en JSON - MODIFIÉ"""
        
        metadata = table_data.get('metadata', {})
        
        # Convertir métadonnées en dict
        if hasattr(metadata, '__dict__'):
            metadata_dict = metadata.__dict__.copy()
        elif hasattr(metadata, '_asdict'):
            metadata_dict = metadata._asdict()
        else:
            metadata_dict = dict(metadata) if metadata else {}
        
        # Ajouter informations du tableau
        metadata_dict.update({
            "title": table_data.get('title', 'Unknown'),
            "headers": table_data.get('headers', []),
            "row_count": len(table_data.get('rows', [])),
            "column_count": len(table_data.get('headers', [])),
            "is_pivot_table": table_data.get('is_pivot', False),
            "processed_records_count": len(table_data.get('processed_records', []))
        })

        with open(metadata_path, "w", encoding="utf-8") as f:
            json.dump(metadata_dict, f, indent=2, ensure_ascii=False)

    def _is_numeric(self, value) -> bool:
        """Détermine si une valeur est numérique"""
        if value is None or value == "":
            return False
        try:
            float(str(value).replace(",", ""))
            return True
        except (ValueError, TypeError):
            return False

    def _generate_filename(self, table_data: Dict[str, Any]) -> str:
        """Génère un nom de fichier basé sur les métadonnées - MODIFIÉ"""
        parts = []
        metadata = table_data.get('metadata', {})
        
        # Extraction sécurisée des métadonnées
        genetic_line = self._safe_get_metadata(metadata, 'genetic_line')
        bird_type = self._safe_get_metadata(metadata, 'bird_type')
        sex = self._safe_get_metadata(metadata, 'sex')
        table_type = self._safe_get_metadata(metadata, 'table_type')

        # Lignée génétique
        if genetic_line and genetic_line != "unknown":
            genetic_clean = genetic_line.replace(" ", "").lower()
            parts.append(genetic_clean)

        # Type de tableau basé sur le titre et métadonnées
        title_lower = table_data.get('title', '').lower()
        
        if table_type and table_type != "unknown":
            parts.append(table_type)
        elif "amino" in title_lower:
            parts.append("amino_acids")
        elif "performance" in title_lower:
            parts.append("performance")
        elif "weight" in title_lower:
            parts.append("weight")
        elif "feed" in title_lower or "nutrition" in title_lower:
            parts.append("nutrition")
        elif "temperature" in title_lower or "environment" in title_lower:
            parts.append("environment")
        else:
            # Utiliser le titre nettoyé
            title_clean = re.sub(r"[^\w\s]", "", title_lower)
            title_clean = re.sub(r"\s+", "_", title_clean)[:20]
            if title_clean and title_clean != "table":
                parts.append(title_clean)

        # Type d'oiseau si pertinent
        if bird_type and bird_type != "unknown" and bird_type != "broiler":
            parts.append(bird_type)

        # Sexe si spécifié
        if sex and sex != "unknown":
            parts.append(sex)

        # Indicateur pivot
        if table_data.get('is_pivot', False):
            parts.append("pivot")

        # Fallback
        if not parts:
            parts.append("table")

        return "_".join(parts)

    def _safe_get_metadata(self, metadata, key: str) -> str:
        """Extraction sécurisée des métadonnées"""
        if hasattr(metadata, key):
            return getattr(metadata, key, "unknown")
        elif isinstance(metadata, dict):
            return metadata.get(key, "unknown")
        else:
            return "unknown"

    def export_summary(self, exported_files: list) -> str:
        """Génère un fichier de résumé de l'extraction"""
        summary_path = self.output_dir / "extraction_summary.json"

        summary = {
            "extraction_date": (
                exported_files[0]["metadata"] if exported_files else None
            ),
            "total_tables": len(exported_files),
            "output_directory": str(self.output_dir),
            "xlsx_support": self.xlsx_available,
            "files": exported_files,
        }

        try:
            with open(summary_path, "w", encoding="utf-8") as f:
                json.dump(summary, f, indent=2, ensure_ascii=False)

            self.logger.info(f"Summary exported: {summary_path}")
            return str(summary_path)

        except Exception as e:
            self.logger.error(f"Error exporting summary: {e}")
            return ""
#!/usr/bin/env python3
"""
Extracteur de tableaux intelligent avec analyse LLM - Version Optimisée
Décomposition des ranges, normalisation pivot et validation renforcée
"""

import json
import csv
import os
import re
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass
from datetime import datetime
import logging
import hashlib
from enum import Enum

# Chargement des variables d'environnement
try:
    from dotenv import load_dotenv

    env_paths = [
        Path(__file__).parent.parent / ".env",
        Path(__file__).parent / ".env",
        Path.cwd() / ".env",
    ]

    for env_path in env_paths:
        if env_path.exists():
            load_dotenv(env_path)
            print(f"Variables d'environnement chargées depuis: {env_path}")
            break
except ImportError:
    print(
        "pip install python-dotenv recommandé pour charger automatiquement les clés API"
    )


class RangeType(Enum):
    """Types de ranges détectés"""

    NUMERIC_RANGE = "numeric_range"  # "3-5"
    EXACT_VALUE = "exact_value"  # "5"
    TEMPORAL_RANGE = "temporal_range"  # "92-98 days"
    CONTEXTUAL_TEXT = "contextual_text"  # "Based on production"
    PERCENTAGE_RANGE = "percentage_range"  # "85-90%"
    EVENT_BASED = "event_based"  # "106-1st Egg"
    HYBRID_RANGE = "hybrid_range"  # "1st Egg-266"
    OPEN_ENDED = "open_ended"  # "> 267"


@dataclass
class DecomposedRange:
    """Range décomposé pour optimisation LLM"""

    original_value: str
    range_type: RangeType

    # Valeurs numériques
    min_value: Optional[float] = None
    max_value: Optional[float] = None
    exact_value: Optional[float] = None

    # Événements biologiques
    min_event: Optional[str] = None
    max_event: Optional[str] = None

    # Métadonnées
    unit: Optional[str] = None
    confidence: float = 1.0
    requires_llm_parsing: bool = False
    is_open_ended: bool = False


@dataclass
class DocumentContext:
    """Contexte d'un document analysé par LLM"""

    genetic_line: str
    document_type: str
    species: str
    measurement_units: str
    target_audience: str
    table_types_expected: List[str]
    confidence_score: float = 0.0
    raw_analysis: str = ""


@dataclass
class TableMetadata:
    """Métadonnées d'un tableau analysées par LLM"""

    table_type: str
    sex: str
    measurement_category: str
    units: str
    age_range: str
    confidence_score: float = 0.0
    reasoning: str = ""


@dataclass
class TableInfo:
    """Information complète sur un tableau extrait"""

    table_id: str
    context: str
    headers: List[str]
    rows: List[List[Any]]
    metadata: Optional[TableMetadata] = None
    document_context: Optional[DocumentContext] = None
    source_file: Optional[str] = None
    extraction_timestamp: Optional[str] = None


class UniversalRangeDecomposer:
    """Décomposeur universel de ranges pour optimisation LLM"""

    def __init__(self):
        # Patterns de reconnaissance
        self.numeric_range_pattern = re.compile(r"^(\d+(?:\.\d+)?)-(\d+(?:\.\d+)?)$")
        self.exact_value_pattern = re.compile(r"^(\d+(?:\.\d+)?)$")
        self.percentage_range_pattern = re.compile(
            r"^(\d+(?:\.\d+)?)-(\d+(?:\.\d+)?)%$"
        )
        self.temporal_pattern = re.compile(r"^(\d+)-(\d+)\s*(days?|weeks?|months?)$")
        self.event_based_pattern = re.compile(
            r"^(\d+)-(1st\s+egg|first\s+egg)$", re.IGNORECASE
        )
        self.hybrid_pattern = re.compile(
            r"^(1st\s+egg|first\s+egg)-(\d+)$", re.IGNORECASE
        )
        self.open_ended_pattern = re.compile(r"^>\s*(\d+)$")

        # Mots-clés contextuels
        self.contextual_keywords = [
            "based on",
            "selon",
            "depending",
            "variable",
            "adjust",
            "monitor",
            "as needed",
            "si nécessaire",
        ]

    def decompose_range(self, value: str, column_context: str = "") -> DecomposedRange:
        """Décompose un range selon son type et contexte"""
        if not value or value.strip() == "":
            return self._create_empty_range(value)

        value = str(value).strip()

        # 1. Range numérique simple : "3-5"
        numeric_match = self.numeric_range_pattern.match(value)
        if numeric_match:
            return DecomposedRange(
                original_value=value,
                range_type=RangeType.NUMERIC_RANGE,
                min_value=float(numeric_match.group(1)),
                max_value=float(numeric_match.group(2)),
                unit=self._extract_unit_from_context(column_context),
                confidence=1.0,
            )

        # 2. Valeur exacte : "5"
        exact_match = self.exact_value_pattern.match(value)
        if exact_match:
            return DecomposedRange(
                original_value=value,
                range_type=RangeType.EXACT_VALUE,
                exact_value=float(exact_match.group(1)),
                unit=self._extract_unit_from_context(column_context),
                confidence=1.0,
            )

        # 3. Range pourcentage : "85-90%"
        percentage_match = self.percentage_range_pattern.match(value)
        if percentage_match:
            return DecomposedRange(
                original_value=value,
                range_type=RangeType.PERCENTAGE_RANGE,
                min_value=float(percentage_match.group(1)),
                max_value=float(percentage_match.group(2)),
                unit="percentage",
                confidence=1.0,
            )

        # 4. Range temporel : "92-98 days"
        temporal_match = self.temporal_pattern.match(value.lower())
        if temporal_match:
            return DecomposedRange(
                original_value=value,
                range_type=RangeType.TEMPORAL_RANGE,
                min_value=float(temporal_match.group(1)),
                max_value=float(temporal_match.group(2)),
                unit=temporal_match.group(3),
                confidence=1.0,
            )

        # 5. Range basé sur événement : "106-1st Egg"
        event_match = self.event_based_pattern.match(value)
        if event_match:
            return DecomposedRange(
                original_value=value,
                range_type=RangeType.EVENT_BASED,
                min_value=float(event_match.group(1)),
                max_event=event_match.group(2).lower().replace(" ", "_"),
                unit="days_to_event",
                confidence=0.8,
                requires_llm_parsing=True,
            )

        # 6. Range hybride : "1st Egg-266"
        hybrid_match = self.hybrid_pattern.match(value)
        if hybrid_match:
            return DecomposedRange(
                original_value=value,
                range_type=RangeType.HYBRID_RANGE,
                min_event=hybrid_match.group(1).lower().replace(" ", "_"),
                max_value=float(hybrid_match.group(2)),
                unit="event_to_days",
                confidence=0.8,
                requires_llm_parsing=True,
            )

        # 7. Range ouvert : "> 267"
        open_match = self.open_ended_pattern.match(value)
        if open_match:
            return DecomposedRange(
                original_value=value,
                range_type=RangeType.OPEN_ENDED,
                min_value=float(open_match.group(1)),
                is_open_ended=True,
                unit="days",
                confidence=0.9,
            )

        # 8. Texte contextuel
        if any(keyword in value.lower() for keyword in self.contextual_keywords):
            return DecomposedRange(
                original_value=value,
                range_type=RangeType.CONTEXTUAL_TEXT,
                requires_llm_parsing=True,
                confidence=0.5,
            )

        # 9. Fallback
        return DecomposedRange(
            original_value=value,
            range_type=RangeType.CONTEXTUAL_TEXT,
            requires_llm_parsing=True,
            confidence=0.3,
        )

    def _extract_unit_from_context(self, column_context: str) -> Optional[str]:
        """Extrait l'unité du contexte de la colonne"""
        if not column_context:
            return None

        unit_patterns = {
            r"g/bird/day": "g_per_bird_per_day",
            r"g/bird": "g_per_bird",
            r"%": "percentage",
            r"°C": "celsius",
            r"days": "days",
            r"weeks": "weeks",
            r"kg": "kg",
            r"lbs": "pounds",
        }

        for pattern, unit in unit_patterns.items():
            if re.search(pattern, column_context, re.IGNORECASE):
                return unit

        return None

    def _create_empty_range(self, value: str) -> DecomposedRange:
        """Crée un range vide pour valeurs manquantes"""
        return DecomposedRange(
            original_value=value, range_type=RangeType.CONTEXTUAL_TEXT, confidence=0.0
        )


class PivotTableNormalizer:
    """Normaliseur pour structures pivot multi-dimensionnelles"""

    def __init__(self, range_decomposer: UniversalRangeDecomposer):
        self.range_decomposer = range_decomposer
        self.logger = logging.getLogger(__name__)

    def is_pivot_table(self, table_info: TableInfo) -> bool:
        """Détermine si un tableau est une structure pivot"""
        headers = [h.lower() for h in table_info.headers]

        # Indicateurs de structure pivot
        phase_indicators = ["starter", "grower", "developer", "breeder", "finisher"]
        age_range_pattern = r"\d+-\d+"

        pivot_score = 0
        for header in headers:
            if any(phase in header for phase in phase_indicators):
                pivot_score += 2
            if re.search(age_range_pattern, header):
                pivot_score += 1
            if any(word in header for word in ["male", "female", "mixed"]):
                pivot_score += 1

        return pivot_score >= 3

    def normalize_pivot_table(self, table_info: TableInfo) -> List[Dict[str, Any]]:
        """Normalise une table pivot en enregistrements individuels"""
        if not self.is_pivot_table(table_info):
            return self._process_standard_table(table_info)

        normalized_records = []

        # Identifier les colonnes de dimension vs valeur
        dimension_cols, value_cols = self._identify_pivot_columns(table_info.headers)

        for row in table_info.rows:
            # Extraire les dimensions communes (nutriment, unité, etc.)
            base_record = {}
            for i, col in enumerate(dimension_cols):
                if i < len(row):
                    base_record[col] = row[i]

            # Créer un enregistrement par colonne de valeur
            for i, value_col in enumerate(value_cols):
                value_idx = len(dimension_cols) + i
                if value_idx < len(row) and row[value_idx]:
                    record = base_record.copy()

                    # Décomposer la phase depuis le header
                    phase_info = self._extract_phase_info(value_col)
                    record.update(phase_info)

                    # Décomposer la valeur
                    value_decomposed = self.range_decomposer.decompose_range(
                        str(row[value_idx]), value_col
                    )
                    record.update(self._range_to_dict(value_decomposed, "value"))

                    normalized_records.append(record)

        return normalized_records

    def _identify_pivot_columns(
        self, headers: List[str]
    ) -> Tuple[List[str], List[str]]:
        """Identifie les colonnes de dimension vs valeur dans un pivot"""
        dimension_cols = []
        value_cols = []

        for header in headers:
            header_lower = header.lower()

            # Colonnes de dimension (descriptives)
            if any(
                dim in header_lower
                for dim in ["phase", "nutrient", "unit", "protein", "energy"]
            ):
                dimension_cols.append(header)
            # Colonnes de valeur (phases/âges)
            elif any(
                phase in header_lower
                for phase in ["starter", "grower", "developer", "breeder", "male"]
            ):
                value_cols.append(header)
            elif re.search(r"\d+-\d+", header):
                value_cols.append(header)
            else:
                dimension_cols.append(header)

        return dimension_cols, value_cols

    def _extract_phase_info(self, column_header: str) -> Dict[str, Any]:
        """Extrait les informations de phase depuis un header de colonne"""
        phase_info = {}
        header_lower = column_header.lower()

        # Phase name
        if "starter" in header_lower:
            phase_info["phase_name"] = "starter"
        elif "grower" in header_lower:
            phase_info["phase_name"] = "grower"
        elif "developer" in header_lower:
            phase_info["phase_name"] = "developer"
        elif "breeder" in header_lower:
            phase_info["phase_name"] = "breeder"
        elif "male" in header_lower:
            phase_info["phase_name"] = "male_specific"
        else:
            phase_info["phase_name"] = "unknown"

        # Age range extraction
        age_range_match = re.search(r"(\d+)-(\d+)", column_header)
        if age_range_match:
            phase_info["age_min_days"] = int(age_range_match.group(1))
            phase_info["age_max_days"] = int(age_range_match.group(2))

        # Event-based ranges
        if "1st egg" in header_lower or "first egg" in header_lower:
            phase_info["event_marker"] = "first_egg"

        # Open-ended ranges
        if ">" in column_header:
            match = re.search(r">\s*(\d+)", column_header)
            if match:
                phase_info["age_min_days"] = int(match.group(1))
                phase_info["is_open_ended"] = True

        return phase_info

    def _process_standard_table(self, table_info: TableInfo) -> List[Dict[str, Any]]:
        """Traite une table standard (non-pivot)"""
        records = []

        for row in table_info.rows:
            record = {}
            for i, header in enumerate(table_info.headers):
                if i < len(row):
                    value = row[i]

                    # Décomposer les ranges dans toutes les colonnes
                    if self._contains_range(str(value)):
                        decomposed = self.range_decomposer.decompose_range(
                            str(value), header
                        )
                        record.update(self._range_to_dict(decomposed, f"col_{i}"))
                    else:
                        record[header] = value

            records.append(record)

        return records

    def _contains_range(self, value: str) -> bool:
        """Détermine si une valeur contient un range à décomposer"""
        return (
            "-" in value
            and re.search(r"\d+-\d+", value)
            or ">" in value
            or any(
                keyword in value.lower()
                for keyword in ["based on", "variable", "depending"]
            )
        )

    def _range_to_dict(
        self, decomposed: DecomposedRange, prefix: str
    ) -> Dict[str, Any]:
        """Convertit un range décomposé en dictionnaire"""
        result = {
            f"{prefix}_original": decomposed.original_value,
            f"{prefix}_type": decomposed.range_type.value,
            f"{prefix}_confidence": decomposed.confidence,
            f"{prefix}_requires_llm": decomposed.requires_llm_parsing,
        }

        if decomposed.min_value is not None:
            result[f"{prefix}_min"] = decomposed.min_value
        if decomposed.max_value is not None:
            result[f"{prefix}_max"] = decomposed.max_value
        if decomposed.exact_value is not None:
            result[f"{prefix}_value"] = decomposed.exact_value
        if decomposed.unit:
            result[f"{prefix}_unit"] = decomposed.unit
        if decomposed.min_event:
            result[f"{prefix}_min_event"] = decomposed.min_event
        if decomposed.max_event:
            result[f"{prefix}_max_event"] = decomposed.max_event
        if decomposed.is_open_ended:
            result[f"{prefix}_open_ended"] = True

        return result


class MetadataEnricher:
    """Enrichisseur de métadonnées basé sur intents.json"""

    def __init__(self, intents_file_path: str = None):
        self.logger = logging.getLogger(__name__)
        self.intents_data = self._load_intents_data(intents_file_path)

    def _load_intents_data(self, intents_file_path: str = None) -> Dict:
        """Charge les données d'intents depuis le fichier JSON"""
        if intents_file_path and Path(intents_file_path).exists():
            try:
                with open(intents_file_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception as e:
                self.logger.warning(f"Erreur chargement intents.json: {e}")

        # Fallback - cherche intents.json dans les répertoires parents
        search_paths = [
            Path(__file__).parent.parent / "intents.json",
            Path(__file__).parent / "intents.json",
            Path.cwd() / "intents.json",
        ]

        for path in search_paths:
            if path.exists():
                try:
                    with open(path, "r", encoding="utf-8") as f:
                        self.logger.info(f"Intents chargé depuis: {path}")
                        return json.load(f)
                except Exception:
                    continue

        self.logger.warning("intents.json non trouvé - métadonnées basiques utilisées")
        return {}

    def normalize_genetic_line(self, raw_line: str) -> str:
        """Normalise une lignée génétique selon les alias d'intents.json"""
        if not self.intents_data or "aliases" not in self.intents_data:
            return raw_line

        line_aliases = self.intents_data["aliases"].get("line", {})
        raw_lower = raw_line.lower().strip()

        for canonical_line, aliases in line_aliases.items():
            if raw_lower == canonical_line.lower():
                return canonical_line
            if raw_lower in [alias.lower() for alias in aliases]:
                return canonical_line

        return raw_line

    def normalize_sex(self, raw_sex: str) -> str:
        """Normalise le sexe selon les alias d'intents.json"""
        if not self.intents_data or "aliases" not in self.intents_data:
            return raw_sex

        sex_aliases = self.intents_data["aliases"].get("sex", {})
        raw_lower = raw_sex.lower().strip()

        for canonical_sex, aliases in sex_aliases.items():
            if raw_lower == canonical_sex.lower():
                return canonical_sex
            if raw_lower in [alias.lower() for alias in aliases]:
                return canonical_sex

        return raw_sex

    def detect_metrics_from_headers(self, headers: List[str]) -> List[str]:
        """Détecte les métriques présentes dans les headers du tableau"""
        if not self.intents_data or "intents" not in self.intents_data:
            return []

        detected_metrics = []
        headers_text = " ".join(headers).lower()

        metric_intent = self.intents_data["intents"].get("metric_query", {})
        metrics_def = metric_intent.get("metrics", {})

        for metric_name, metric_config in metrics_def.items():
            search_terms = self._metric_to_search_terms(metric_name)
            if any(term in headers_text for term in search_terms):
                detected_metrics.append(metric_name)

        return detected_metrics

    def _metric_to_search_terms(self, metric_name: str) -> List[str]:
        """Convertit un nom de métrique en termes de recherche"""
        metric_mapping = {
            "body_weight_target": ["body weight", "weight", "poids", "live weight"],
            "body_weight_avg": ["body weight", "weight", "poids", "average weight"],
            "daily_gain": ["daily gain", "gain", "croissance", "daily growth"],
            "fcr_target": ["fcr", "feed conversion", "conversion ratio", "ic"],
            "uniformity_pct": ["uniformity", "uniformité", "cv"],
            "mortality_expected_pct": ["mortality", "mortalité", "survival"],
            "feed_intake_daily": ["feed intake", "consommation", "aliment"],
            "water_intake_daily": ["water intake", "eau", "water consumption"],
            "egg_production_pct": ["production", "ponte", "egg"],
            "ambient_temp_target": ["temperature", "température", "temp"],
            "humidity_target": ["humidity", "humidité", "rh"],
            "lighting_hours": ["lighting", "light", "éclairage", "photopériode"],
            "me_kcalkg": ["energy", "me", "kcal", "énergie"],
            "cp_pct": ["protein", "protéine", "cp", "crude protein"],
            "lys_digestible_pct": ["lysine", "lys"],
            "ca_pct": ["calcium", "ca"],
            "av_p_pct": ["phosphorus", "phosphore", "available p"],
        }

        return metric_mapping.get(metric_name, [metric_name.replace("_", " ")])


class LLMClient:
    """Interface pour différents clients LLM avec chargement automatique des clés"""

    def __init__(self, provider="openai", api_key=None, model="gpt-4"):
        self.provider = provider.lower()

        # Chargement automatique de la clé API depuis .env
        if api_key is None:
            if self.provider == "openai":
                api_key = os.getenv("OPENAI_API_KEY")
            elif self.provider == "anthropic":
                api_key = os.getenv("ANTHROPIC_API_KEY")

        self.api_key = api_key
        self.model = model
        self._setup_client()

    def _setup_client(self):
        """Configure le client LLM selon le provider"""
        if self.provider == "openai":
            if not self.api_key:
                raise ValueError("Clé OPENAI_API_KEY manquante dans .env ou paramètres")
            try:
                import openai

                self.client = openai.OpenAI(api_key=self.api_key)
                self.logger = logging.getLogger(__name__)
                self.logger.info(f"Client OpenAI configuré avec modèle {self.model}")
            except ImportError:
                raise ImportError("pip install openai required for OpenAI provider")

        elif self.provider == "anthropic":
            if not self.api_key:
                raise ValueError(
                    "Clé ANTHROPIC_API_KEY manquante dans .env ou paramètres"
                )
            try:
                import anthropic

                self.client = anthropic.Anthropic(api_key=self.api_key)
                self.logger = logging.getLogger(__name__)
                self.logger.info(f"Client Anthropic configuré avec modèle {self.model}")
            except ImportError:
                raise ImportError(
                    "pip install anthropic required for Anthropic provider"
                )

        elif self.provider == "mock":
            self.client = None
        else:
            raise ValueError(f"Provider non supporté: {self.provider}")

    @classmethod
    def create_auto(cls, provider="openai", model="gpt-4"):
        """Factory method pour créer un client avec chargement automatique de clé"""
        return cls(provider=provider, model=model)

    def complete(self, prompt: str, max_tokens: int = 1000) -> str:
        """Complète un prompt avec le LLM"""
        if self.provider == "mock":
            return self._mock_response(prompt)

        elif self.provider == "openai":
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=max_tokens,
                temperature=0.1,
            )
            return response.choices[0].message.content

        elif self.provider == "anthropic":
            response = self.client.messages.create(
                model=self.model,
                max_tokens=max_tokens,
                temperature=0.1,
                messages=[{"role": "user", "content": prompt}],
            )
            return response.content[0].text

    def _mock_response(self, prompt: str) -> str:
        """Réponse mock pour tests"""
        if "analyze this document" in prompt.lower():
            return json.dumps(
                {
                    "genetic_line": "Ross 308",
                    "document_type": "nutrition_specifications",
                    "species": "broilers",
                    "measurement_units": "metric",
                    "target_audience": "nutritionists",
                    "table_types_expected": [
                        "amino_acid_profile",
                        "feed_specifications",
                        "particle_size",
                    ],
                    "confidence_score": 0.9,
                }
            )
        elif "analyze this table" in prompt.lower():
            return json.dumps(
                {
                    "table_type": "amino_acid_profile",
                    "sex": "mixed",
                    "measurement_category": "nutrition",
                    "units": "percentage",
                    "age_range": "0-52_days",
                    "confidence_score": 0.8,
                    "reasoning": "Table shows amino acid ratios for different age periods",
                }
            )
        return "{}"


class DocumentAnalyzer:
    """Analyseur de documents avec LLM"""

    def __init__(self, llm_client: LLMClient):
        self.llm_client = llm_client
        self.logger = logging.getLogger(__name__)

    def analyze_document(self, json_file: str, txt_file: str = None) -> DocumentContext:
        """Analyse un document pour extraire le contexte global"""
        try:
            sample = self._extract_document_sample(json_file, txt_file)
            prompt = self._build_document_analysis_prompt(sample, json_file)
            response = self.llm_client.complete(prompt, max_tokens=800)
            context = self._parse_document_response(response)

            self.logger.info(
                f"Document analysé: {context.genetic_line} - {context.document_type}"
            )
            return context

        except Exception as e:
            self.logger.error(f"Erreur analyse document: {e}")
            return self._fallback_document_context(json_file)

    def _extract_document_sample(self, json_file: str, txt_file: str = None) -> str:
        """Extrait un échantillon représentatif du document"""
        sample_parts = []

        try:
            with open(json_file, "r", encoding="utf-8") as f:
                json_data = json.load(f)

            if "metadata" in json_data:
                sample_parts.append(
                    f"Metadata: {json.dumps(json_data['metadata'], indent=2)}"
                )

            if "text" in json_data:
                text_sample = json_data["text"][:2000]
                sample_parts.append(f"Text sample: {text_sample}")

            if "chunks" in json_data and isinstance(json_data["chunks"], list):
                for i, chunk in enumerate(json_data["chunks"][:3]):
                    if isinstance(chunk, str):
                        chunk_sample = chunk[:500]
                        sample_parts.append(f"Chunk {i}: {chunk_sample}")

        except Exception as e:
            sample_parts.append(f"JSON error: {e}")

        if txt_file and Path(txt_file).exists():
            try:
                with open(txt_file, "r", encoding="utf-8") as f:
                    txt_sample = f.read(1500)
                    sample_parts.append(f"TXT sample: {txt_sample}")
            except Exception as e:
                sample_parts.append(f"TXT error: {e}")

        return "\n\n".join(sample_parts)

    def _build_document_analysis_prompt(self, sample: str, filename: str) -> str:
        """Construit le prompt d'analyse du document"""
        return f"""
Analyze this agricultural/poultry document and extract key metadata.

Filename: {Path(filename).name}

Document sample:
{sample}

Return ONLY a valid JSON object with these fields:
{{
    "genetic_line": "exact genetic line (Ross 308, Ross 708, Cobb 500, Cobb 700, Hubbard Classic, etc.) or Unknown",
    "document_type": "type (performance_objectives, nutrition_specifications, management_guide, handbook, feed_guide, etc.)",
    "species": "target species (broilers, layers, breeders, turkeys, etc.)",
    "measurement_units": "primary units (metric, imperial, mixed)",
    "target_audience": "primary audience (producers, nutritionists, veterinarians, managers, etc.)",
    "table_types_expected": ["list", "of", "expected", "table", "types"],
    "confidence_score": 0.0-1.0
}}

Table types can include: performance_data, nutrition_specifications, amino_acid_profile, feed_form, carcass_yield, management_guidelines, vitamin_mineral, feed_ingredients, etc.

Focus on extracting precise, specific information. If uncertain, use descriptive terms rather than guessing.
"""

    def _parse_document_response(self, response: str) -> DocumentContext:
        """Parse et valide la réponse LLM"""
        try:
            response = response.strip()
            if response.startswith("```json"):
                response = response[7:]
            if response.endswith("```"):
                response = response[:-3]

            data = json.loads(response)

            return DocumentContext(
                genetic_line=data.get("genetic_line", "Unknown"),
                document_type=data.get("document_type", "unknown"),
                species=data.get("species", "unknown"),
                measurement_units=data.get("measurement_units", "unknown"),
                target_audience=data.get("target_audience", "unknown"),
                table_types_expected=data.get("table_types_expected", []),
                confidence_score=float(data.get("confidence_score", 0.0)),
                raw_analysis=response,
            )

        except (json.JSONDecodeError, KeyError, ValueError) as e:
            self.logger.error(f"Erreur parsing réponse LLM: {e}")
            raise

    def _fallback_document_context(self, json_file: str) -> DocumentContext:
        """Contexte de fallback en cas d'échec LLM"""
        filename = Path(json_file).name.lower()

        genetic_line = "Unknown"
        if "ross308" in filename or "ross_308" in filename:
            genetic_line = "Ross 308"
        elif "cobb500" in filename:
            genetic_line = "Cobb 500"

        return DocumentContext(
            genetic_line=genetic_line,
            document_type="unknown",
            species="broilers",
            measurement_units="unknown",
            target_audience="unknown",
            table_types_expected=[],
            confidence_score=0.3,
        )


class TableAnalyzer:
    """Analyseur de tableaux avec LLM et enrichissement métadonnées"""

    def __init__(
        self, llm_client: LLMClient, metadata_enricher: MetadataEnricher = None
    ):
        self.llm_client = llm_client
        self.metadata_enricher = metadata_enricher
        self.logger = logging.getLogger(__name__)

    def analyze_table(
        self, table_info: TableInfo, document_context: DocumentContext
    ) -> TableMetadata:
        """Analyse un tableau pour extraire ses métadonnées"""
        try:
            prompt = self._build_table_analysis_prompt(table_info, document_context)
            response = self.llm_client.complete(prompt, max_tokens=600)
            metadata = self._parse_table_response(response)

            self.logger.info(f"Tableau analysé: {metadata.table_type} - {metadata.sex}")
            return metadata

        except Exception as e:
            self.logger.error(f"Erreur analyse tableau: {e}")
            return self._fallback_table_metadata(table_info, document_context)

    def _build_table_analysis_prompt(
        self, table_info: TableInfo, document_context: DocumentContext
    ) -> str:
        """Construit le prompt d'analyse du tableau"""
        sample_rows = (
            table_info.rows[:5] if len(table_info.rows) > 5 else table_info.rows
        )

        return f"""
Analyze this table from a {document_context.genetic_line} {document_context.document_type} document.

Document context:
- Genetic line: {document_context.genetic_line}
- Document type: {document_context.document_type}
- Species: {document_context.species}
- Expected table types: {document_context.table_types_expected}

Table information:
- Context/Title: {table_info.context}
- Headers: {table_info.headers}
- Sample rows: {sample_rows}
- Total rows: {len(table_info.rows)}

Return ONLY a valid JSON object:
{{
    "table_type": "specific table type (performance_data, amino_acid_profile, feed_form, carcass_yield, etc.)",
    "sex": "target sex (male, female, as-hatched, mixed, unknown)",
    "measurement_category": "category (performance, nutrition, management, processing, etc.)",
    "units": "primary units (metric, imperial, percentage, ratio, etc.)",
    "age_range": "age range (0-10_days, 11-24_days, 0-52_days, adult, etc.)",
    "confidence_score": 0.0-1.0,
    "reasoning": "brief explanation of analysis"
}}

Unit determination rules:
- "metric" if weight data in grams (g), kilograms (kg), or metric measurements
- "imperial" if weight data in pounds (lb, lbs), ounces (oz), or imperial measurements  
- "percentage" for ratios, percentages, or proportional data
- "ratio" for amino acid ratios or feed conversion ratios

Sex determination rules:
- "male" if specifically for male birds
- "female" if specifically for female birds  
- "as-hatched" if mixed-sex performance data
- "mixed" if nutritional/management data applying to both sexes
- "unknown" if cannot be determined

Be precise and specific in your analysis, especially for unit detection.
"""

    def _parse_table_response(self, response: str) -> TableMetadata:
        """Parse et valide la réponse LLM pour un tableau"""
        try:
            response = response.strip()
            if response.startswith("```json"):
                response = response[7:]
            if response.endswith("```"):
                response = response[:-3]

            data = json.loads(response)

            return TableMetadata(
                table_type=data.get("table_type", "unknown"),
                sex=data.get("sex", "unknown"),
                measurement_category=data.get("measurement_category", "unknown"),
                units=data.get("units", "unknown"),
                age_range=data.get("age_range", "unknown"),
                confidence_score=float(data.get("confidence_score", 0.0)),
                reasoning=data.get("reasoning", ""),
            )

        except (json.JSONDecodeError, KeyError, ValueError) as e:
            self.logger.error(f"Erreur parsing réponse tableau LLM: {e}")
            raise

    def _fallback_table_metadata(
        self, table_info: TableInfo, document_context: DocumentContext
    ) -> TableMetadata:
        """Métadonnées de fallback en cas d'échec LLM"""
        context_lower = table_info.context.lower()

        if "performance" in context_lower:
            table_type = "performance_data"
            measurement_category = "performance"
        elif "amino acid" in context_lower:
            table_type = "amino_acid_profile"
            measurement_category = "nutrition"
        elif "carcass" in context_lower:
            table_type = "carcass_yield"
            measurement_category = "processing"
        else:
            table_type = "unknown"
            measurement_category = "unknown"

        return TableMetadata(
            table_type=table_type,
            sex="unknown",
            measurement_category=measurement_category,
            units="unknown",
            age_range="unknown",
            confidence_score=0.3,
            reasoning="Fallback analysis",
        )


class BaseTableExtractor:
    """Extracteur de base pour la détection structurelle des tableaux"""

    def __init__(self):
        self.logger = logging.getLogger(__name__)

    def extract_tables_from_files(
        self, json_file: str, txt_file: str = None
    ) -> List[TableInfo]:
        """Extrait les tableaux des fichiers JSON/TXT"""
        timestamp = datetime.now().isoformat()

        try:
            with open(json_file, "r", encoding="utf-8") as f:
                json_data = json.load(f)

            txt_content = None
            if txt_file and Path(txt_file).exists():
                with open(txt_file, "r", encoding="utf-8") as f:
                    txt_content = f.read()

            if txt_content:
                tables = self._extract_markdown_tables(
                    txt_content, json_file, timestamp
                )
            else:
                tables = self._extract_from_json_only(json_data, json_file, timestamp)

            return tables

        except Exception as e:
            self.logger.error(f"Erreur extraction: {e}")
            return []

    def _extract_markdown_tables(
        self, text: str, source_file: str, timestamp: str
    ) -> List[TableInfo]:
        """Extrait les tableaux Markdown du texte"""
        tables = []
        lines = text.split("\n")

        i = 0
        table_count = 0

        while i < len(lines):
            line = lines[i].strip()

            if line.startswith("|") and line.endswith("|") and line.count("|") >= 3:
                context = self._find_table_context(lines, i)
                table_lines, end_index = self._extract_markdown_table_lines(lines, i)

                if len(table_lines) >= 3:
                    headers, rows = self._parse_markdown_table(table_lines)

                    if headers and rows and len(rows) > 2:
                        table_count += 1
                        table_id = (
                            f"table_{table_count:03d}_{self._clean_context(context)}"
                        )

                        table_info = TableInfo(
                            table_id=table_id,
                            context=context,
                            headers=headers,
                            rows=rows,
                            source_file=source_file,
                            extraction_timestamp=timestamp,
                        )

                        tables.append(table_info)

                i = end_index
            else:
                i += 1

        return tables

    def _extract_from_json_only(
        self, json_data: Dict, source_file: str, timestamp: str
    ) -> List[TableInfo]:
        """Extrait les tableaux depuis le JSON uniquement"""
        tables = []

        if "text" in json_data:
            tables.extend(
                self._extract_markdown_tables(json_data["text"], source_file, timestamp)
            )

        if "chunks" in json_data and isinstance(json_data["chunks"], list):
            for chunk_idx, chunk in enumerate(json_data["chunks"]):
                if isinstance(chunk, str):
                    chunk_tables = self._extract_markdown_tables(
                        chunk, source_file, timestamp
                    )
                    tables.extend(chunk_tables)

        return tables

    def _find_table_context(self, lines: List[str], table_start_index: int) -> str:
        """Trouve le contexte d'un tableau"""
        for i in range(table_start_index - 1, max(0, table_start_index - 10), -1):
            line = lines[i].strip()

            if not line or line.startswith("---") or line.startswith("*"):
                continue

            if line.startswith("##"):
                return re.sub(r"^#+\s*", "", line).strip()

            if line and not line.startswith("|") and len(line) < 80:
                return re.sub(r"[#*_`]", "", line).strip()

        return "unknown_table"

    def _extract_markdown_table_lines(
        self, lines: List[str], start_index: int
    ) -> Tuple[List[str], int]:
        """Extrait les lignes d'un tableau Markdown"""
        table_lines = []
        i = start_index

        while i < len(lines):
            line = lines[i].strip()

            if line.startswith("|") and line.endswith("|"):
                table_lines.append(line)
                i += 1
            elif not line:
                i += 1
                j = i
                while j < len(lines) and not lines[j].strip():
                    j += 1
                if j < len(lines) and lines[j].strip().startswith("|"):
                    continue
                else:
                    break
            else:
                break

        return table_lines, i

    def _parse_markdown_table(
        self, table_lines: List[str]
    ) -> Tuple[List[str], List[List[str]]]:
        """Parse les lignes d'un tableau Markdown"""
        if len(table_lines) < 2:
            return [], []

        headers = self._parse_markdown_table_row(table_lines[0])

        if not headers:
            return [], []

        rows = []
        for line in table_lines[2:]:
            if line.strip():
                row = self._parse_markdown_table_row(line)
                if row:
                    while len(row) < len(headers):
                        row.append("")
                    rows.append(row[: len(headers)])

        return headers, rows

    def _parse_markdown_table_row(self, line: str) -> List[str]:
        """Parse une ligne de tableau Markdown"""
        line = line.strip()
        if line.startswith("|"):
            line = line[1:]
        if line.endswith("|"):
            line = line[:-1]

        cells = []
        for cell in line.split("|"):
            cell = cell.strip()
            cell = re.sub(r"[\\*_`]", "", cell)
            cell = re.sub(r"\s+", " ", cell).strip()
            cells.append(cell)

        return cells

    def _clean_context(self, context: str) -> str:
        """Nettoie un contexte pour nom de fichier"""
        clean = re.sub(r"[^\w\s-]", "", context.lower())
        clean = re.sub(r"[-\s]+", "_", clean)
        return clean[:30] if clean else "table"


class DataValidator:
    """Validateur pour cohérence et qualité des données extraites"""

    def __init__(self):
        self.logger = logging.getLogger(__name__)

    def validate_record(self, record: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """Valide un enregistrement individuel"""
        errors = []

        # Validation âge/semaine cohérence
        if "age_min_days" in record and "age_max_days" in record:
            min_days = record.get("age_min_days")
            max_days = record.get("age_max_days")

            if min_days and max_days:
                try:
                    min_days, max_days = float(min_days), float(max_days)
                    if min_days >= max_days:
                        errors.append(f"Age range invalide: {min_days} >= {max_days}")
                    if min_days < 0 or max_days > 1000:
                        errors.append(
                            f"Age hors limite réaliste: {min_days}-{max_days}"
                        )
                except (ValueError, TypeError):
                    errors.append("Ages non numériques")

        # Validation ranges numériques
        for key in record:
            if key.endswith("_min") and f"{key[:-4]}_max" in record:
                try:
                    min_val = float(record[key])
                    max_val = float(record[f"{key[:-4]}_max"])
                    if min_val >= max_val:
                        errors.append(f"Range invalide {key}: {min_val} >= {max_val}")
                except (ValueError, TypeError):
                    continue

        # Validation métadonnées essentielles
        required_fields = ["genetic_line", "sex", "table_type"]
        for field in required_fields:
            if not record.get(field) or record.get(field) == "unknown":
                errors.append(f"Champ essentiel manquant: {field}")

        return len(errors) == 0, errors

    def validate_dataset(self, records: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Valide un jeu de données complet"""
        total_records = len(records)
        valid_records = 0
        all_errors = []

        for i, record in enumerate(records):
            is_valid, errors = self.validate_record(record)
            if is_valid:
                valid_records += 1
            else:
                all_errors.extend([f"Record {i+1}: {error}" for error in errors])

        validation_result = {
            "total_records": total_records,
            "valid_records": valid_records,
            "error_rate": (
                (total_records - valid_records) / total_records
                if total_records > 0
                else 0
            ),
            "errors": all_errors[:20],  # Limite pour éviter overflow
            "quality_score": valid_records / total_records if total_records > 0 else 0,
        }

        self.logger.info(
            f"Validation: {valid_records}/{total_records} valides "
            f"(score: {validation_result['quality_score']:.2f})"
        )

        return validation_result


class OptimizedTableExtractor:
    """Extracteur de tableaux optimisé avec décomposition ranges et normalisation pivot"""

    def __init__(
        self,
        output_dir: str = "extracted_tables",
        llm_client: LLMClient = None,
        intents_file: str = None,
    ):
        """Initialise l'extracteur optimisé"""
        # Configuration de base avec répertoire suggéré
        if not os.path.isabs(output_dir):
            # Répertoire suggéré pour PerformanceMetrics
            suggested_base = Path(
                r"C:\intelia_gpt\intelia-expert\rag\documents\PerformanceMetrics"
            )
            self.output_dir = suggested_base / output_dir
        else:
            self.output_dir = Path(output_dir)

        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Configuration du logging
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)

        # Client LLM
        if llm_client is None:
            self.llm_client = LLMClient(provider="mock")
        else:
            self.llm_client = llm_client

        # Composants spécialisés
        self.metadata_enricher = MetadataEnricher(intents_file)
        self.range_decomposer = UniversalRangeDecomposer()
        self.pivot_normalizer = PivotTableNormalizer(self.range_decomposer)
        self.validator = DataValidator()

        # Analyseurs
        self.document_analyzer = DocumentAnalyzer(self.llm_client)
        self.table_analyzer = TableAnalyzer(self.llm_client, self.metadata_enricher)

        # Extracteur de base
        self.base_extractor = BaseTableExtractor()

        self.logger.info(
            f"Extracteur optimisé initialisé - LLM: {self.llm_client.provider}"
        )

    def process_document(self, json_file: str, txt_file: str = None) -> List[str]:
        """Traite un document avec décomposition avancée et normalisation"""
        try:
            self.logger.info(f"Traitement optimisé: {Path(json_file).name}")

            # 1. Analyse du contexte du document
            document_context = self.document_analyzer.analyze_document(
                json_file, txt_file
            )

            # 2. Normalisation avec intents.json
            document_context.genetic_line = (
                self.metadata_enricher.normalize_genetic_line(
                    document_context.genetic_line
                )
            )

            # 3. Extraction structurelle des tableaux
            raw_tables = self.base_extractor.extract_tables_from_files(
                json_file, txt_file
            )

            if not raw_tables:
                self.logger.warning(f"Aucun tableau trouvé dans {json_file}")
                return []

            self.logger.info(f"Tableaux détectés: {len(raw_tables)}")

            # 4. Analyse LLM enrichie de chaque tableau
            all_records = []
            for table in raw_tables:
                table.document_context = document_context
                table.metadata = self.table_analyzer.analyze_table(
                    table, document_context
                )

                # Normalisation des métadonnées avec intents.json
                if table.metadata:
                    table.metadata.sex = self.metadata_enricher.normalize_sex(
                        table.metadata.sex
                    )

                # NOUVEAU: Décomposition avancée avec normalisation pivot
                normalized_records = self.pivot_normalizer.normalize_pivot_table(table)

                # Enrichissement avec métadonnées du document
                for record in normalized_records:
                    record.update(self._build_document_metadata(document_context))
                    record.update(self._build_table_metadata(table.metadata))

                all_records.extend(normalized_records)

            # 5. Validation des données
            validation_result = self.validator.validate_dataset(all_records)
            self.logger.info(
                f"Qualité des données: {validation_result['quality_score']:.2f}"
            )

            # 6. Export optimisé
            csv_files = self._export_normalized_records(
                all_records, document_context, Path(json_file).stem
            )

            self.logger.info(f"Export terminé: {len(csv_files)} fichiers générés")
            return csv_files

        except Exception as e:
            self.logger.error(f"Erreur traitement document: {e}")
            return []

    def _build_document_metadata(self, context: DocumentContext) -> Dict[str, Any]:
        """Construit les métadonnées du document"""
        return {
            "genetic_line": context.genetic_line,
            "document_type": context.document_type,
            "species": context.species,
            "measurement_units": context.measurement_units,
            "target_audience": context.target_audience,
        }

    def _build_table_metadata(
        self, metadata: Optional[TableMetadata]
    ) -> Dict[str, Any]:
        """Construit les métadonnées du tableau avec décomposition âge"""
        if not metadata:
            return {
                "sex": "unknown",
                "table_type": "unknown",
                "measurement_category": "unknown",
            }

        # Décomposition de l'age_range
        age_decomposed = self.range_decomposer.decompose_range(
            metadata.age_range, "age range"
        )

        result = {
            "sex": metadata.sex,
            "table_type": metadata.table_type,
            "measurement_category": metadata.measurement_category,
            "units_system": metadata.units,
        }

        # Ajout des âges décomposés
        if age_decomposed.min_value is not None:
            result["age_min_days"] = int(age_decomposed.min_value)
        if age_decomposed.max_value is not None:
            result["age_max_days"] = int(age_decomposed.max_value)
        if age_decomposed.min_value and age_decomposed.max_value:
            result["age_min_weeks"] = max(1, int(age_decomposed.min_value) // 7)
            result["age_max_weeks"] = int(age_decomposed.max_value) // 7

        return result

    def _export_normalized_records(
        self, records: List[Dict[str, Any]], context: DocumentContext, base_name: str
    ) -> List[str]:
        """Exporte les enregistrements normalisés"""
        if not records:
            return []

        # Génération nom de fichier intelligent
        filename = self._generate_optimized_filename(context, base_name)
        csv_path = self.output_dir / f"{filename}.csv"
        xlsx_path = self.output_dir / f"{filename}.xlsx"

        try:
            # Récupération de toutes les clés pour headers
            all_keys = set()
            for record in records:
                all_keys.update(record.keys())

            headers = sorted(list(all_keys))

            # Export CSV
            with open(csv_path, "w", newline="", encoding="utf-8-sig") as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=headers)
                writer.writeheader()
                writer.writerows(records)

            # Export XLSX avec formatage
            self._export_to_xlsx_optimized(xlsx_path, records, headers, context)

            return [str(csv_path)]

        except Exception as e:
            self.logger.error(f"Erreur export: {e}")
            return []

    def _export_to_xlsx_optimized(
        self,
        xlsx_path: Path,
        records: List[Dict],
        headers: List[str],
        context: DocumentContext,
    ):
        """Export XLSX optimisé avec formatage avancé"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            self.logger.warning("openpyxl non installé - skip export XLSX")
            return

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Normalized_Data"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        metadata_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Headers
        metadata_cols = [
            "genetic_line",
            "document_type",
            "species",
            "sex",
            "table_type",
            "measurement_category",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.border = border

            if header in metadata_cols:
                cell.fill = metadata_fill
            else:
                cell.fill = header_fill

        # Données
        for row_idx, record in enumerate(records, 2):
            for col_idx, header in enumerate(headers, 1):
                value = record.get(header, "")
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border

                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")

        # Auto-ajustement colonnes
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(xlsx_path)

    def _generate_optimized_filename(
        self, context: DocumentContext, base_name: str
    ) -> str:
        """Génère un nom de fichier optimisé"""
        parts = []

        if context.genetic_line != "Unknown":
            genetic_clean = context.genetic_line.replace(" ", "_").lower()
            parts.append(genetic_clean)

        if context.document_type != "unknown":
            parts.append(context.document_type)

        parts.append("normalized")

        # Hash pour unicité
        content_hash = hashlib.md5(
            f"{base_name}_{context.genetic_line}_{context.document_type}".encode()
        ).hexdigest()[:6]
        parts.append(content_hash)

        filename = "_".join(parts)
        filename = re.sub(r"[^\w\-_.]", "", filename)

        return filename

    def process_directory(
        self, json_dir: str, pattern: str = "*.json", recursive: bool = True
    ) -> Dict[str, List[str]]:
        """Traite tous les fichiers d'un répertoire avec optimisations"""
        json_dir_path = Path(json_dir)

        if not json_dir_path.exists():
            raise FileNotFoundError(f"Répertoire non trouvé: {json_dir}")

        if recursive:
            json_files = list(json_dir_path.rglob(pattern))
        else:
            json_files = list(json_dir_path.glob(pattern))

        if not json_files:
            self.logger.warning(f"Aucun fichier JSON trouvé dans {json_dir}")
            return {}

        self.logger.info(f"Traitement optimisé de {len(json_files)} fichiers")

        results = {}
        for json_file in json_files:
            try:
                txt_file = self._find_corresponding_txt_file(str(json_file))
                csv_files = self.process_document(str(json_file), txt_file)
                results[str(json_file)] = csv_files

            except Exception as e:
                self.logger.error(f"Erreur avec {json_file.name}: {e}")
                results[str(json_file)] = []

        return results

    def _find_corresponding_txt_file(self, json_file_path: str) -> str:
        """Trouve le fichier TXT correspondant au fichier JSON"""
        json_path = Path(json_file_path)

        possible_txt_paths = [
            json_path.with_suffix(".txt"),
            json_path.parent
            / (json_path.name.replace("_extracted.json", "_extracted.txt")),
            json_path.parent / (json_path.stem.replace("_extracted", "") + ".txt"),
        ]

        for txt_path in possible_txt_paths:
            if txt_path.exists():
                return str(txt_path)

        return None


def main():
    """Interface en ligne de commande optimisée"""
    print("=== EXTRACTEUR DE TABLEAUX OPTIMISÉ ===\n")

    # Configuration LLM
    use_llm = input("Utiliser un vrai LLM ? [O/n]: ").strip().lower()
    use_real_llm = use_llm not in ["n", "no", "non"]

    if use_real_llm:
        provider = (
            input("Provider LLM (openai/anthropic) [openai]: ").strip() or "openai"
        )

        try:
            llm_client = LLMClient.create_auto(provider=provider)
            print("Clé API chargée automatiquement depuis .env")
        except ValueError as e:
            print(f"Erreur: {e}")
            api_key = input("Clé API (fallback): ").strip()
            if not api_key:
                print("Passage en mode MOCK")
                llm_client = LLMClient(provider="mock")
            else:
                llm_client = LLMClient(provider=provider, api_key=api_key)
        except Exception as e:
            print(f"Erreur configuration LLM: {e}")
            print("Passage en mode MOCK")
            llm_client = LLMClient(provider="mock")
    else:
        print("Mode MOCK activé - analyse basique")
        llm_client = LLMClient(provider="mock")

    # Chemin d'entrée
    input_path = input("Chemin vers vos fichiers JSON: ").strip()
    if not input_path or not Path(input_path).exists():
        print("Chemin invalide")
        return 1

    # Répertoire de sortie optionnel
    output_dir = (
        input("Répertoire de sortie [extracted_tables_optimized]: ").strip()
        or "extracted_tables_optimized"
    )

    # Initialisation de l'extracteur optimisé
    extractor = OptimizedTableExtractor(output_dir=output_dir, llm_client=llm_client)

    try:
        path_obj = Path(input_path)

        if path_obj.is_file():
            txt_file = extractor._find_corresponding_txt_file(str(path_obj))
            csv_files = extractor.process_document(str(path_obj), txt_file)

            print(f"\n{len(csv_files)} fichier(s) généré(s)")
            for csv_file in csv_files:
                print(f"   {Path(csv_file).name}")

        elif path_obj.is_dir():
            results = extractor.process_directory(str(path_obj))
            total_files = sum(len(csv_files) for csv_files in results.values())
            successful_files = sum(1 for csv_files in results.values() if csv_files)

            print("\nRÉSUMÉ:")
            print(f"   Fichiers traités: {len(results)}")
            print(f"   Fichiers réussis: {successful_files}")
            print(f"   Fichiers générés: {total_files}")

            for json_file, csv_files in results.items():
                filename = Path(json_file).name
                if csv_files:
                    print(f"   {filename}: {len(csv_files)} fichier(s)")
                else:
                    print(f"   {filename}: Aucun fichier généré")

        print(f"\nFichiers sauvegardés dans: {extractor.output_dir}")

        if use_real_llm and llm_client.provider != "mock":
            print(f"\nLLM utilisé: {provider} - Analyse intelligente + Optimisations")
        else:
            print("\nMode MOCK - Optimisations structurelles actives")

    except Exception as e:
        print(f"Erreur: {e}")
        return 1

    input("\nAppuyez sur Entrée pour fermer...")
    return 0


def quick_extract_optimized(
    json_path: str,
    output_dir: str = "extracted_tables_optimized",
    use_real_llm: bool = True,
    provider: str = "openai",
):
    """Fonction utilitaire pour extraction rapide optimisée"""
    if use_real_llm:
        try:
            llm_client = LLMClient.create_auto(provider=provider)
        except Exception:
            print("Clé API non trouvée, utilisation du mode MOCK")
            llm_client = LLMClient(provider="mock")
    else:
        llm_client = LLMClient(provider="mock")

    extractor = OptimizedTableExtractor(output_dir=output_dir, llm_client=llm_client)
    path_obj = Path(json_path)

    if path_obj.is_file():
        txt_file = extractor._find_corresponding_txt_file(str(path_obj))
        csv_files = extractor.process_document(str(path_obj), txt_file)
        return len(csv_files)
    elif path_obj.is_dir():
        results = extractor.process_directory(str(path_obj))
        return sum(len(csv_files) for csv_files in results.values())
    else:
        raise FileNotFoundError(f"Chemin invalide: {json_path}")


# Exemple d'utilisation des nouvelles fonctionnalités
def demo_optimized_features():
    """Démonstration des nouvelles fonctionnalités"""

    # 1. Décomposition de ranges
    decomposer = UniversalRangeDecomposer()

    test_ranges = [
        ("3-5", "Feed Increase Range (g/bird/day)"),
        ("106-1st Egg", "Developer phase"),
        ("> 267", "Breeder 2 phase"),
        ("Based on production", "Variable feeding"),
        ("85-90%", "Uniformity target"),
    ]

    print("=== DÉMONSTRATION DÉCOMPOSITION RANGES ===")
    for value, context in test_ranges:
        decomposed = decomposer.decompose_range(value, context)
        print(f"\nOriginal: '{value}'")
        print(f"Type: {decomposed.range_type.value}")
        print(f"Min: {decomposed.min_value}, Max: {decomposed.max_value}")
        print(f"Unité: {decomposed.unit}")
        print(f"Nécessite LLM: {decomposed.requires_llm_parsing}")

    # 2. Validation de données
    validator = DataValidator()

    test_record = {
        "genetic_line": "Cobb500",
        "sex": "female",
        "table_type": "nutrition_specifications",
        "age_min_days": 92,
        "age_max_days": 126,
        "feed_min": 3,
        "feed_max": 5,
    }

    print("\n=== DÉMONSTRATION VALIDATION ===")
    is_valid, errors = validator.validate_record(test_record)
    print(f"Enregistrement valide: {is_valid}")
    if errors:
        print(f"Erreurs: {errors}")

    return True


if __name__ == "__main__":
    # Démonstration optionnelle des nouvelles fonctionnalités
    if len(os.sys.argv) > 1 and os.sys.argv[1] == "--demo":
        demo_optimized_features()
    else:
        exit(main())

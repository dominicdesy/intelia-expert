#!/usr/bin/env python3
"""
PDF Table Extractor using Claude Vision API
============================================

Utilise l'API Claude avec vision pour extraire intelligemment les tableaux
de n'importe quel PDF, sans dépendre de mots-clés spécifiques.

Usage:
    python extract_pdf_tables_claude_vision.py <pdf_file> <output_excel>

Example:
    python extract_pdf_tables_claude_vision.py Cobb500-2022.pdf output.xlsx
"""

import anthropic
import sys
import json
import base64
from pathlib import Path
from typing import List, Dict
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import fitz  # PyMuPDF
from PIL import Image
from io import BytesIO
import os


class ClaudeVisionTableExtractor:
    """Extracteur de tableaux utilisant Claude Vision API"""

    def __init__(self, pdf_path: str, api_key: str = None):
        self.pdf_path = Path(pdf_path)
        self.api_key = api_key or os.getenv('ANTHROPIC_API_KEY')

        if not self.api_key:
            raise ValueError("ANTHROPIC_API_KEY non trouvé. Définir la variable d'environnement ou passer api_key")

        self.client = anthropic.Anthropic(api_key=self.api_key)
        self.tables = []

    def convert_pdf_to_images(self, dpi: int = 300) -> List:
        """Convertit le PDF en images en utilisant PyMuPDF"""
        print(f"[OK] Conversion du PDF en images (DPI={dpi})...")

        pdf_document = fitz.open(str(self.pdf_path))
        images = []

        # Convertir chaque page en image
        zoom = dpi / 72  # 72 DPI par défaut
        mat = fitz.Matrix(zoom, zoom)

        for page_num in range(len(pdf_document)):
            page = pdf_document[page_num]
            pix = page.get_pixmap(matrix=mat)

            # Convertir en PIL Image
            img_data = pix.tobytes("png")
            image = Image.open(BytesIO(img_data))
            images.append(image)

        pdf_document.close()

        print(f"  {len(images)} pages converties")
        return images

    def image_to_base64(self, image) -> str:
        """Convertit une image PIL en base64"""
        buffered = BytesIO()
        image.save(buffered, format="PNG")
        return base64.b64encode(buffered.getvalue()).decode('utf-8')

    def extract_table_from_image(self, image, page_num: int) -> Dict:
        """
        Extrait le tableau d'une image de page en utilisant Claude Vision

        Returns:
            Dict avec la structure:
            {
                'page': int,
                'table_found': bool,
                'table_title': str,
                'table_type': str,  # 'performance', 'nutrition', 'amino_acids', 'yield'
                'metadata': dict,
                'headers': list,
                'data': list of lists
            }
        """
        print(f"\n[Processing] Page {page_num}...")

        # Convertir l'image en base64
        image_base64 = self.image_to_base64(image)

        # Prompt pour Claude
        prompt = """Analyse cette page de PDF et extrait TOUS les tableaux de données présents.

Pour chaque tableau trouvé, retourne un objet JSON avec cette structure EXACTE:

{
  "tables": [
    {
      "table_number": "numéro du tableau si indiqué (ex: 'Table 1', 'Table 2', '1', '2', etc.), null sinon",
      "table_title": "titre exact du tableau (ex: 'C500 Broiler Performance Objectives (Metric) - Female')",
      "table_type": "type de tableau (ex: 'performance', 'nutrition', 'amino_acids', 'yield', 'other')",
      "description": "description du contenu du tableau",
      "headers": ["colonne1", "colonne2", "colonne3", ...],
      "units": ["unité1", "unité2", "unité3", ...],
      "data": [
        ["valeur1", "valeur2", "valeur3", ...],
        ["valeur1", "valeur2", "valeur3", ...],
        ...
      ],
      "metadata": {
        "sex": "mixed/male/female (si applicable)",
        "unit_system": "metric/imperial (si applicable)",
        "bird_type": "type d'animal (si applicable)",
        "age_range": "plage d'âge (si applicable)"
      }
    }
  ]
}

IMPORTANT:
- Extrait TOUTES les lignes de données, pas seulement un échantillon
- Préserve l'ordre exact des colonnes
- Inclus les en-têtes avec leurs unités
- Si un numéro de table est visible (ex: "Table 1", "Table 2"), capture-le dans "table_number"
- Si une page contient plusieurs tableaux, retourne-les tous dans le tableau "tables"
- Si aucun tableau n'est trouvé, retourne: {"tables": []}
- Retourne UNIQUEMENT le JSON, sans texte avant ou après"""

        try:
            # Appel à Claude Vision API
            message = self.client.messages.create(
                model="claude-3-5-sonnet-20250318",  # Latest model with vision
                max_tokens=4096,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": "image/png",
                                    "data": image_base64,
                                },
                            },
                            {
                                "type": "text",
                                "text": prompt
                            }
                        ],
                    }
                ],
            )

            # Parser la réponse JSON
            response_text = message.content[0].text

            # Extraire le JSON (enlever les possibles markdown code blocks)
            if "```json" in response_text:
                response_text = response_text.split("```json")[1].split("```")[0].strip()
            elif "```" in response_text:
                response_text = response_text.split("```")[1].split("```")[0].strip()

            result = json.loads(response_text)

            # Ajouter le numéro de page
            for table in result.get('tables', []):
                table['page'] = page_num

            print(f"  [OK] {len(result.get('tables', []))} tableau(x) trouvé(s)")

            return result

        except json.JSONDecodeError as e:
            print(f"  [ERROR] Erreur de parsing JSON: {e}")
            print(f"  Response: {response_text[:500]}...")
            return {'tables': []}
        except Exception as e:
            print(f"  [ERROR] Erreur lors de l'extraction: {e}")
            return {'tables': []}

    def extract_all_tables(self, start_page: int = 1, end_page: int = None) -> List[Dict]:
        """
        Extrait tous les tableaux du PDF

        Args:
            start_page: Première page à traiter (1-indexed)
            end_page: Dernière page à traiter (1-indexed), None = toutes
        """
        # Convertir le PDF en images
        images = self.convert_pdf_to_images()

        # Ajuster les indices (conversion 1-indexed vers 0-indexed)
        start_idx = max(0, start_page - 1)
        end_idx = min(len(images), end_page) if end_page else len(images)

        all_tables = []

        print(f"\n{'='*60}")
        print(f"Extraction des pages {start_page} à {end_idx}")
        print(f"{'='*60}")

        for i in range(start_idx, end_idx):
            page_num = i + 1
            result = self.extract_table_from_image(images[i], page_num)

            for table in result.get('tables', []):
                all_tables.append(table)

        print(f"\n{'='*60}")
        print(f"[OK] Total: {len(all_tables)} tableau(x) extrait(s)")

        # Fusionner les tables multi-pages
        all_tables = self._merge_multi_page_tables(all_tables)
        print(f"[OK] Après fusion multi-pages: {len(all_tables)} tableau(x)")
        print(f"{'='*60}")

        return all_tables

    def _merge_multi_page_tables(self, tables: List[Dict]) -> List[Dict]:
        """
        Fusionne automatiquement les tables qui s'étendent sur plusieurs pages.
        Utilise table_number pour identifier les tables à fusionner.
        """
        if not tables:
            return tables

        # Grouper les tables par table_number
        table_groups = {}
        tables_without_number = []

        for table in tables:
            table_num = table.get('table_number')

            if not table_num or table_num == '' or table_num is None:
                # Pas de numéro, garder tel quel
                tables_without_number.append(table)
            else:
                # Nettoyer le numéro (enlever "Table " si présent)
                clean_num = str(table_num).replace('Table', '').replace('table', '').strip()

                if clean_num not in table_groups:
                    table_groups[clean_num] = []
                table_groups[clean_num].append(table)

        # Fusionner les tables du même groupe
        merged_tables = []

        for table_num, group in table_groups.items():
            if len(group) == 1:
                # Une seule page, pas de fusion nécessaire
                merged_tables.append(group[0])
            else:
                # Multiple pages - fusionner
                print(f"  [Fusion] Table {table_num}: {len(group)} pages détectées")
                merged_table = self._merge_table_group(group)
                merged_tables.append(merged_table)

        # Ajouter les tables sans numéro
        merged_tables.extend(tables_without_number)

        # Trier par page source
        merged_tables.sort(key=lambda t: t.get('page', 0))

        return merged_tables

    def _merge_table_group(self, tables: List[Dict]) -> Dict:
        """
        Fusionne un groupe de tables (même table sur plusieurs pages)
        """
        if not tables:
            return {}

        # Trier par page
        tables = sorted(tables, key=lambda t: t.get('page', 0))

        # Prendre les métadonnées de la première page
        merged = tables[0].copy()

        # Fusionner les données
        all_data = []
        for table in tables:
            data = table.get('data', [])
            all_data.extend(data)

        merged['data'] = all_data

        # Ajouter une note sur les pages sources
        pages = [t.get('page') for t in tables if t.get('page')]
        if len(pages) > 1:
            merged['source_pages'] = f"{min(pages)}-{max(pages)}"
            merged['page_count'] = len(pages)

        return merged

    def save_to_excel(self, tables: List[Dict], output_path: str):
        """Sauvegarde tous les tableaux dans un fichier Excel avec métadonnées"""
        print(f"\n{'='*60}")
        print(f"Création du fichier Excel: {output_path}")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Supprimer la feuille par défaut

        for idx, table in enumerate(tables):
            # Générer un nom de feuille (max 31 caractères pour Excel)
            sheet_name = self._generate_sheet_name(table, idx)

            print(f"  -> Création feuille: {sheet_name}")

            # Créer le DataFrame
            df = pd.DataFrame(table['data'], columns=table['headers'])

            # Préparer les métadonnées
            metadata = self._prepare_metadata(table)

            # Informations sur les colonnes
            column_info = self._infer_column_types(table['headers'], table.get('units', []), df)

            # Créer la feuille Excel
            self.create_excel_sheet(wb, sheet_name, metadata, df, column_info)

        # Sauvegarder
        wb.save(output_path)
        print(f"\n[OK] Fichier Excel sauvegarde: {output_path}")
        print(f"  Feuilles creees: {len(wb.sheetnames)}")

    def _generate_sheet_name(self, table: Dict, index: int) -> str:
        """Génère un nom de feuille valide et unique"""
        # Extraire des infos du titre ou métadonnées
        title = table.get('table_title', '')
        table_type = table.get('table_type', 'table')
        table_number = table.get('table_number', '')
        metadata = table.get('metadata', {})

        # Nettoyer le numéro de table
        clean_table_num = str(table_number).replace('Table', '').replace('table', '').strip() if table_number else ''

        # Construire un nom basé sur le type et les métadonnées
        parts = []

        # Si on a un numéro de table, le préfixer
        if clean_table_num:
            parts.append(f'T{clean_table_num}')

        if 'performance' in title.lower() or table_type == 'performance':
            sex = metadata.get('sex', 'mixed')
            unit_system = metadata.get('unit_system', 'metric')
            parts.extend([sex, unit_system])
        elif 'nutrition' in title.lower() or 'nutrient' in title.lower():
            if 'small' in title.lower():
                parts.extend(['nutrient', 'small'])
            else:
                parts.extend(['nutrient', 'med_large'])
        elif 'amino' in title.lower():
            if 'small' in title.lower():
                parts.extend(['amino', 'small'])
            else:
                parts.extend(['amino', 'med_large'])
        elif 'yield' in title.lower():
            sex = metadata.get('sex', 'mixed')
            parts.extend(['yield', sex, 'metric'])
        else:
            # Si pas de numéro, utiliser l'index
            if not clean_table_num:
                parts = [table_type, str(index + 1)]
            else:
                parts.append(table_type)

        sheet_name = '_'.join(parts)

        # Limiter à 31 caractères (limite Excel)
        return sheet_name[:31]

    def _prepare_metadata(self, table: Dict) -> Dict:
        """Prépare les métadonnées pour le fichier Excel"""
        metadata = {}

        # Pages sources (peut être une seule page ou une plage)
        if 'source_pages' in table:
            metadata['source_pages'] = table['source_pages']
            metadata['page_count'] = table.get('page_count', 1)
        else:
            metadata['source_page'] = table.get('page', '')

        # Informations de table
        metadata['table_number'] = table.get('table_number', '')
        metadata['table_title'] = table.get('table_title', '')
        metadata['table_type'] = table.get('table_type', '')
        metadata['description'] = table.get('description', '')

        # Ajouter les métadonnées spécifiques du tableau
        table_metadata = table.get('metadata', {})
        metadata.update(table_metadata)

        return metadata

    def _infer_column_types(self, headers: List[str], units: List[str], df: pd.DataFrame) -> List[Dict]:
        """Infère les types de colonnes"""
        column_info = []

        for i, header in enumerate(headers):
            unit = units[i] if i < len(units) else ''

            # Nettoyer le nom
            col_name = str(header).replace(' ', '_').replace('(', '').replace(')', '').replace('.', '')

            # Déterminer le type
            col_type = 'text'
            if i < len(df.columns):
                # Analyser les valeurs de la colonne
                col_values = df.iloc[:, i].dropna()
                if len(col_values) > 0:
                    numeric_count = sum(1 for v in col_values if str(v).replace('.', '').replace(',', '').replace('-', '').isdigit())
                    if numeric_count / len(col_values) > 0.8:
                        # Si plus de 80% sont des nombres
                        has_decimal = any('.' in str(v) for v in col_values)
                        col_type = 'numeric' if has_decimal else 'integer'

            column_info.append({
                'name': col_name,
                'type': col_type,
                'unit': unit or 'various'
            })

        return column_info

    def create_excel_sheet(self, workbook: openpyxl.Workbook, sheet_name: str,
                          metadata: Dict, df: pd.DataFrame, column_info: List[Dict]):
        """Crée une feuille Excel avec métadonnées et données"""
        ws = workbook.create_sheet(sheet_name)

        # Styles
        metadata_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        metadata_header_font = Font(color="FFFFFF", bold=True)
        data_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        data_header_font = Font(bold=True)

        # Écrire les métadonnées
        row = 1
        ws.cell(row, 1, 'metadata').font = metadata_header_font
        ws.cell(row, 1).fill = metadata_header_fill
        ws.cell(row, 2, 'value').font = metadata_header_font
        ws.cell(row, 2).fill = metadata_header_fill
        row += 1

        for key, value in metadata.items():
            ws.cell(row, 1, key)
            ws.cell(row, 2, value)
            row += 1

        # Informations de table
        ws.cell(row, 1, 'table_data_rows')
        ws.cell(row, 2, len(df))
        row += 1
        ws.cell(row, 1, 'table_columns')
        ws.cell(row, 2, len(df.columns))
        row += 1

        # Informations sur les colonnes
        for i, col_info in enumerate(column_info, 1):
            ws.cell(row, 1, f'column_{i}_name')
            ws.cell(row, 2, col_info['name'])
            row += 1
            ws.cell(row, 1, f'column_{i}_type')
            ws.cell(row, 2, col_info['type'])
            row += 1
            ws.cell(row, 1, f'column_{i}_unit')
            ws.cell(row, 2, col_info['unit'])
            row += 1

        # Lignes vides
        row += 2

        # En-tête des données
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row, col_idx, col_name)
            cell.fill = data_header_fill
            cell.font = data_header_font
            cell.alignment = Alignment(horizontal='center')
        row += 1

        # Données
        for _, data_row in df.iterrows():
            for col_idx, value in enumerate(data_row, 1):
                ws.cell(row, col_idx, value)
            row += 1

        # Ajuster largeur des colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width


def main():
    """Fonction principale"""
    if len(sys.argv) < 3:
        print("Usage: python extract_pdf_tables_claude_vision.py <pdf_file> <output_excel> [start_page] [end_page]")
        print("\nExample:")
        print("  python extract_pdf_tables_claude_vision.py Cobb500-2022.pdf output.xlsx")
        print("  python extract_pdf_tables_claude_vision.py Cobb500-2022.pdf output.xlsx 4 6")
        sys.exit(1)

    pdf_file = sys.argv[1]
    output_excel = sys.argv[2]
    start_page = int(sys.argv[3]) if len(sys.argv) > 3 else 1
    end_page = int(sys.argv[4]) if len(sys.argv) > 4 else None

    if not Path(pdf_file).exists():
        print(f"[ERROR] Le fichier {pdf_file} n'existe pas")
        sys.exit(1)

    print("="*60)
    print("PDF Table Extractor - Claude Vision API")
    print("="*60)

    try:
        extractor = ClaudeVisionTableExtractor(pdf_file)

        # Extraire les tableaux
        tables = extractor.extract_all_tables(start_page=start_page, end_page=end_page)

        if not tables:
            print("[ERROR] Aucun tableau trouve dans le PDF")
            sys.exit(1)

        # Sauvegarder en Excel
        extractor.save_to_excel(tables, output_excel)

        print("\n" + "="*60)
        print("[OK] Extraction terminee avec succes!")
        print("="*60)

    except Exception as e:
        print(f"\n[ERROR] {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()

#!/usr/bin/env python3
"""
Extracteur de données Excel intelligent
"""

import re
import logging
from typing import List, Tuple, Optional, Dict, Any

from openpyxl.worksheet.worksheet import Worksheet
import openpyxl

from models import MetricData
from config import IntentsConfigLoader

logger = logging.getLogger(__name__)


class IntelligentDataExtractor:
    """Extracteur universel avec reconnaissance de métriques via intents.json"""

    def __init__(
        self,
        format_type: str,
        data_type: str,
        intents_config: Optional[IntentsConfigLoader] = None,
    ):
        self.format_type = format_type
        self.data_type = data_type
        self.intents_config = intents_config or IntentsConfigLoader()
        self.metric_aliases = self.intents_config.get_metric_types()

        self.category_mapping = {
            "performance": ["performance", "growth", "weight", "gain", "fcr"]
            + self.metric_aliases.get("performance", []),
            "nutrition": ["nutrition", "feed", "amino", "vitamin", "mineral", "protein"]
            + self.metric_aliases.get("consumption", []),
            "pharmaceutical": ["drug", "vaccine", "medication", "treatment", "dosage"],
            "carcass": ["carcass", "yield", "processing", "parts", "weight"],
            "environment": ["temperature", "humidity", "light", "space", "housing"]
            + self.metric_aliases.get("environment", []),
            "health": ["mortality", "disease", "symptom", "diagnosis"],
            "economics": ["cost", "price", "profit", "economics"]
            + self.metric_aliases.get("economics", []),
            "other": [],
        }

    def extract_metrics(self, workbook: openpyxl.Workbook) -> List[MetricData]:
        """Extraction adaptative avec reconnaissance intelligente"""
        all_metrics = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            if self.data_type == "performance":
                metrics = self._extract_performance_data(sheet, sheet_name)
            elif self.data_type == "pharmaceutical":
                metrics = self._extract_pharmaceutical_data(sheet, sheet_name)
            elif self.data_type == "nutrition":
                metrics = self._extract_nutrition_data(sheet, sheet_name)
            elif self.data_type == "carcass":
                metrics = self._extract_carcass_data(sheet, sheet_name)
            else:
                metrics = self._extract_generic_data(sheet, sheet_name)

            all_metrics.extend(metrics)

        logger.info(
            f"Total métriques extraites: {len(all_metrics)} (type: {self.data_type})"
        )
        return all_metrics

    def _extract_performance_data(
        self, sheet: Worksheet, sheet_name: str
    ) -> List[MetricData]:
        """Extraction spécialisée pour données de performance"""
        if self._has_metadata_format(sheet):
            return self._extract_metadata_sheet(sheet, sheet_name)
        else:
            return self._extract_tabular_performance_data(sheet, sheet_name)

    def _extract_metadata_sheet(
        self, sheet: Worksheet, sheet_name: str
    ) -> List[MetricData]:
        """Extraction format metadata/value avec support métadonnées de structure"""
        table_metadata = self._extract_table_metadata(sheet)

        if table_metadata:
            logger.info(
                f"Métadonnées de table trouvées pour {sheet_name}: {table_metadata}"
            )
            return self._extract_with_metadata_guidance(
                sheet, sheet_name, table_metadata
            )
        else:
            logger.warning(
                f"Pas de métadonnées de structure pour {sheet_name}, extraction classique"
            )
            return self._extract_classic_metadata(sheet, sheet_name)

    def _extract_table_metadata(self, sheet: Worksheet) -> Optional[Dict[str, Any]]:
        """Extrait les métadonnées de structure de table"""
        metadata = {}

        for row in sheet.iter_rows(min_row=1, max_row=50):
            if not row[0].value or not row[1].value:
                continue

            key = str(row[0].value).strip().lower()
            value = str(row[1].value).strip()

            if key == "table_header_row":
                metadata["header_row"] = int(value)
            elif key == "table_data_rows":
                metadata["data_rows"] = int(value)
            elif key == "table_columns":
                metadata["columns"] = int(value)
            elif key == "expected_metrics":
                metadata["expected_metrics"] = int(value)
            elif key == "validation_checksum":
                metadata["checksum"] = value
            elif key.startswith("column_") and key.endswith("_name"):
                col_num = key.split("_")[1]
                if "column_definitions" not in metadata:
                    metadata["column_definitions"] = {}
                if col_num not in metadata["column_definitions"]:
                    metadata["column_definitions"][col_num] = {}
                metadata["column_definitions"][col_num]["name"] = value
            elif key.startswith("column_") and key.endswith("_type"):
                col_num = key.split("_")[1]
                if "column_definitions" not in metadata:
                    metadata["column_definitions"] = {}
                if col_num not in metadata["column_definitions"]:
                    metadata["column_definitions"][col_num] = {}
                metadata["column_definitions"][col_num]["type"] = value
            elif key.startswith("column_") and key.endswith("_unit"):
                col_num = key.split("_")[1]
                if "column_definitions" not in metadata:
                    metadata["column_definitions"] = {}
                if col_num not in metadata["column_definitions"]:
                    metadata["column_definitions"][col_num] = {}
                metadata["column_definitions"][col_num]["unit"] = value
            else:
                if key.startswith("metadata_"):
                    if "descriptive_metadata" not in metadata:
                        metadata["descriptive_metadata"] = {}
                    metadata["descriptive_metadata"][key] = value
                elif key in [
                    "housing_system",
                    "bird_type",
                    "site_type",
                    "sex",
                    "name",
                    "description",
                ]:
                    normalized_key = f"metadata_{key}"
                    if "descriptive_metadata" not in metadata:
                        metadata["descriptive_metadata"] = {}
                    metadata["descriptive_metadata"][normalized_key] = value
                elif key not in ["brand", "breed", "strain", "type", "species"]:
                    if "descriptive_metadata" not in metadata:
                        metadata["descriptive_metadata"] = {}
                    metadata["descriptive_metadata"][key] = value

        total_metadata_count = len(metadata.get("descriptive_metadata", {}))
        logger.info(f"Métadonnées descriptives capturées: {total_metadata_count}")
        if total_metadata_count > 0:
            logger.debug(
                f"Clés métadonnées: {list(metadata.get('descriptive_metadata', {}).keys())}"
            )

        if "header_row" in metadata and "data_rows" in metadata:
            logger.info(f"Métadonnées de structure détectées: {metadata}")
            return metadata

        return None

    def detect_unit_system(self, table_metadata: Optional[Dict[str, Any]] = None) -> str:
        """
        Détecte le système d'unités dominant du document (metric/imperial/mixed)

        Args:
            table_metadata: Métadonnées de table contenant les définitions de colonnes

        Returns:
            str: 'metric', 'imperial', ou 'mixed'
        """
        if not table_metadata or "column_definitions" not in table_metadata:
            # Par défaut, assume metric (standard international aviculture)
            logger.debug("Pas de métadonnées de colonnes, assume metric par défaut")
            return "metric"

        units = []
        for col_def in table_metadata["column_definitions"].values():
            if "unit" in col_def:
                unit_str = str(col_def["unit"]).lower()
                units.append(unit_str)

        if not units:
            logger.debug("Aucune unité trouvée dans les métadonnées, assume metric")
            return "metric"

        # Listes d'unités par système
        metric_units = ['grams', 'kilograms', 'kg', 'g', 'cm', 'mm', 'meter', 'celsius', '°c']
        imperial_units = ['pounds', 'lb', 'lbs', 'oz', 'ounce', 'ounces', 'inches', 'inch', 'feet', 'foot', 'fahrenheit', '°f']

        # Compter les unités de chaque système
        metric_count = sum(1 for u in units if any(metric_unit in u for metric_unit in metric_units))
        imperial_count = sum(1 for u in units if any(imperial_unit in u for imperial_unit in imperial_units))

        # Déterminer le système dominant
        if metric_count > 0 and imperial_count > 0:
            result = "mixed"
            logger.info(f"Système d'unités mixte détecté (metric: {metric_count}, imperial: {imperial_count})")
        elif imperial_count > 0:
            result = "imperial"
            logger.info(f"Système d'unités impérial détecté ({imperial_count} unités)")
        else:
            result = "metric"
            logger.info(f"Système d'unités métrique détecté ({metric_count} unités)")

        logger.debug(f"Unités trouvées: {units}")
        return result

    def _extract_with_metadata_guidance(
        self, sheet: Worksheet, sheet_name: str, table_metadata: Dict[str, Any]
    ) -> List[MetricData]:
        """Extraction dirigée par métadonnées"""
        metrics = []
        category = self._categorize_sheet(sheet_name)
        self._current_table_metadata = table_metadata

        header_row = table_metadata["header_row"]
        data_rows = table_metadata["data_rows"]
        expected_metrics = table_metadata.get("expected_metrics", 0)
        checksum = table_metadata.get("checksum", "")

        logger.info(
            f"Extraction dirigée: en-têtes ligne {header_row}, {data_rows} lignes"
        )

        headers = []
        header_row_data = sheet[header_row]
        for cell in header_row_data:
            if cell.value and str(cell.value).strip():
                headers.append(str(cell.value).strip())
            else:
                break

        logger.info(f"En-têtes extraits: {headers}")

        first_col_type = self._detect_first_column_type(
            sheet, header_row, headers[0] if headers else ""
        )
        logger.info(f"Type de première colonne détecté: {first_col_type}")

        for row_idx in range(header_row + 1, header_row + 1 + data_rows):
            try:
                row = sheet[row_idx]
                first_cell = row[0]
                if first_cell.value is None or first_cell.value == "":
                    logger.warning(f"Ligne {row_idx}: première cellule vide, arrêt")
                    break

                first_value = str(first_cell.value).strip()

                if first_col_type == "age_days":
                    try:
                        age = int(float(first_value))
                        primary_key = f"day_{age}"
                        age_min, age_max = age, age
                    except (ValueError, TypeError):
                        logger.warning(
                            f"Ligne {row_idx}: âge invalide '{first_value}', ignoré"
                        )
                        continue
                elif first_col_type == "weight_grams":
                    try:
                        weight = int(float(first_value))
                        primary_key = f"weight_{weight}g"
                        age_min, age_max = None, None
                    except (ValueError, TypeError):
                        logger.warning(
                            f"Ligne {row_idx}: poids invalide '{first_value}', ignoré"
                        )
                        continue
                elif first_col_type == "text_label":
                    primary_key = (
                        first_value.lower()
                        .replace(" ", "_")
                        .replace("(", "")
                        .replace(")", "")
                        .replace("%", "pct")
                    )
                    age_min, age_max = None, None
                else:
                    primary_key = first_value.lower().replace(" ", "_")
                    age_min, age_max = None, None

                for col_idx, header in enumerate(headers):
                    if col_idx >= len(row) or not row[col_idx].value:
                        continue

                    cell = row[col_idx]

                    if col_idx == 0:
                        metric = self._create_primary_key_metric(
                            sheet_name,
                            category,
                            first_col_type,
                            cell,
                            header,
                            primary_key,
                            age_min,
                            age_max,
                            checksum,
                            row_idx,
                            col_idx,
                        )
                        metrics.append(metric)
                        continue

                    normalized_metric = self._normalize_metric_name(header)
                    metric_key = f"{primary_key}_{normalized_metric}"
                    metric_name = f"{normalized_metric} for {first_value}"

                    value_numeric, unit = self._parse_numeric_value(str(cell.value))

                    if "column_definitions" in table_metadata:
                        col_def = table_metadata["column_definitions"].get(
                            str(col_idx + 1), {}
                        )
                        if "unit" in col_def and not unit:
                            unit = col_def["unit"]

                    metric = MetricData(
                        sheet_name=sheet_name,
                        category=category,
                        metric_key=metric_key,
                        metric_name=metric_name,
                        value_text=str(cell.value),
                        value_numeric=value_numeric,
                        unit=unit,
                        age_min=age_min,
                        age_max=age_max,
                        metadata={
                            "format": "metadata_guided",
                            "data_type": self.data_type,
                            "original_header": header,
                            "normalized_name": normalized_metric,
                            "primary_key": primary_key,
                            "column_type": first_col_type,
                            "checksum": checksum,
                            "row": row_idx,
                            "col": col_idx,
                        },
                    )

                    metrics.append(metric)

            except Exception as e:
                logger.error(f"Erreur ligne {row_idx}: {e}")
                continue

        actual_metrics = len(metrics)
        if expected_metrics > 0:
            if actual_metrics == expected_metrics:
                logger.info(f"✅ Validation réussie: {actual_metrics} métriques")
            else:
                logger.warning(
                    f"⚠️ Validation échouée: {actual_metrics} vs {expected_metrics}"
                )

        logger.info(
            f"Extraction terminée: {actual_metrics} métriques pour {sheet_name}"
        )
        return metrics

    def _create_primary_key_metric(
        self,
        sheet_name: str,
        category: str,
        first_col_type: str,
        cell,
        header: str,
        primary_key: str,
        age_min: Optional[int],
        age_max: Optional[int],
        checksum: str,
        row_idx: int,
        col_idx: int,
    ) -> MetricData:
        """Crée une métrique de clé primaire selon le type"""
        base_metadata = {
            "format": "metadata_guided",
            "data_type": self.data_type,
            "original_header": header,
            "column_type": first_col_type,
            "checksum": checksum,
            "row": row_idx,
            "col": col_idx,
            "is_primary_key": True,
        }

        if first_col_type == "age_days":
            return MetricData(
                sheet_name=sheet_name,
                category=category,
                metric_key=f"age_{primary_key}",
                metric_name=f"Age at {primary_key}",
                value_text=str(cell.value),
                value_numeric=float(age_min) if age_min is not None else None,
                unit="days",
                age_min=age_min,
                age_max=age_max,
                metadata=base_metadata,
            )
        elif first_col_type == "weight_grams":
            weight_value = str(cell.value).strip()
            return MetricData(
                sheet_name=sheet_name,
                category=category,
                metric_key=f"reference_{primary_key}",
                metric_name=f"Live weight: {weight_value}g",
                value_text=str(cell.value),
                value_numeric=(
                    float(weight_value)
                    if weight_value.replace(".", "", 1).isdigit()
                    else None
                ),
                unit="grams",
                age_min=age_min,
                age_max=age_max,
                metadata=base_metadata,
            )
        else:
            return MetricData(
                sheet_name=sheet_name,
                category=category,
                metric_key=f"label_{primary_key}",
                metric_name=f"Nutrient/Item: {str(cell.value)}",
                value_text=str(cell.value),
                metadata=base_metadata,
            )

    def _detect_first_column_type(
        self, sheet: Worksheet, header_row: int, first_header: str
    ) -> str:
        """Détecte le type de données dans la première colonne"""
        header_lower = first_header.lower()

        if any(
            keyword in header_lower
            for keyword in ["age", "day", "jour", "week", "semaine"]
        ):
            return "age_days"

        if any(
            keyword in header_lower
            for keyword in ["weight", "poids", "live weight", "poids vif"]
        ):
            return "weight_grams"

        if any(
            keyword in header_lower
            for keyword in ["nutrient", "amino", "acid", "protein", "acide"]
        ):
            return "text_label"

        sample_values = []
        for row_idx in range(header_row + 1, min(header_row + 6, sheet.max_row + 1)):
            cell = sheet.cell(row_idx, 1)
            if cell.value:
                sample_values.append(str(cell.value).strip())

        if not sample_values:
            return "unknown"

        numeric_count = 0
        text_count = 0

        for value in sample_values:
            try:
                float(value)
                numeric_count += 1
            except ValueError:
                text_count += 1

        if text_count > numeric_count:
            return "text_label"

        try:
            first_num = float(sample_values[0])
            if 0 <= first_num <= 600:
                return "age_days"
            elif 500 <= first_num <= 10000:
                return "weight_grams"
        except ValueError:
            pass

        return "text_label"

    def _extract_classic_metadata(
        self, sheet: Worksheet, sheet_name: str
    ) -> List[MetricData]:
        """Méthode d'extraction classique (fallback)"""
        metrics = []
        category = self._categorize_sheet(sheet_name)

        for row in sheet.iter_rows(min_row=2):
            if not row[0].value:
                break

            metric_key = str(row[0].value).strip()
            value_raw = row[1].value if row[1].value else ""

            if metric_key.lower().startswith(("table_", "column_", "validation_")):
                continue

            if metric_key.lower() in [
                "brand",
                "breed",
                "strain",
                "type",
                "species",
                "housing_system",
                "feather_color",
                "sex",
            ]:
                continue

            value_numeric, unit = self._parse_numeric_value(str(value_raw))
            age_min, age_max = self._parse_age_range(metric_key, str(value_raw))

            metric = MetricData(
                sheet_name=sheet_name,
                category=category,
                metric_key=metric_key,
                metric_name=self._clean_metric_name(metric_key),
                value_text=str(value_raw) if value_raw else None,
                value_numeric=value_numeric,
                unit=unit,
                age_min=age_min,
                age_max=age_max,
                metadata={"format": "classic_metadata", "data_type": self.data_type},
            )

            metrics.append(metric)

        return metrics

    def _extract_tabular_performance_data(
        self, sheet: Worksheet, sheet_name: str
    ) -> List[MetricData]:
        """Extraction format tabulaire"""
        metrics = []
        category = "performance"

        header_row, headers = self._find_headers_row(sheet)
        if not headers:
            return metrics

        logger.info(f"En-têtes trouvés ligne {header_row}: {headers}")

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row = sheet[row_idx]
            age_cell = row[0]
            if not age_cell.value:
                break

            try:
                age = int(float(age_cell.value))
            except (ValueError, TypeError):
                logger.warning(
                    f"Âge invalide ligne {row_idx}: '{age_cell.value}', ignoré"
                )
                continue

            for col_idx, cell in enumerate(row):
                if col_idx >= len(headers) or not cell.value:
                    continue

                if col_idx == 0:
                    metric = MetricData(
                        sheet_name=sheet_name,
                        category=category,
                        metric_key=f"age_day_{age}",
                        metric_name=f"Age at day {age}",
                        value_text=str(cell.value),
                        value_numeric=float(age),
                        unit="days",
                        age_min=age,
                        age_max=age,
                        metadata={
                            "format": "tabular",
                            "row": row_idx,
                            "col": col_idx,
                            "is_age_column": True,
                            "original_header": headers[col_idx],
                        },
                    )
                    metrics.append(metric)
                    continue

                normalized_metric = self._normalize_metric_name(headers[col_idx])
                metric_key = f"day_{age}_{normalized_metric}"
                metric_name = f"{normalized_metric} at day {age}"

                value_numeric, unit = self._parse_numeric_value(str(cell.value))

                metric = MetricData(
                    sheet_name=sheet_name,
                    category=category,
                    metric_key=metric_key,
                    metric_name=metric_name,
                    value_text=str(cell.value),
                    value_numeric=value_numeric,
                    unit=unit,
                    age_min=age,
                    age_max=age,
                    metadata={
                        "format": "tabular_performance",
                        "data_type": "performance",
                        "original_header": headers[col_idx],
                        "normalized_name": normalized_metric,
                        "row": row_idx,
                        "col": col_idx,
                    },
                )

                metrics.append(metric)

        logger.info(
            f"Extraction tabulaire terminée: {len(metrics)} métriques pour {sheet_name}"
        )
        return metrics

    def _normalize_metric_name(self, header: str) -> str:
        """Normalise le nom d'une métrique"""
        header_lower = header.lower()

        metric_mappings = {
            "weight_g": "body_weight",
            "weight": "body_weight",
            "poids": "body_weight",
            "daily_gain_g": "daily_gain",
            "daily_gain": "daily_gain",
            "gain": "daily_gain",
            "av_daily_gain_g": "average_daily_gain",
            "daily_intake_g": "feed_intake",
            "intake": "feed_intake",
            "cum_intake_g": "cumulative_feed_intake",
            "fcr": "feed_conversion_ratio",
            "feed_conversion": "feed_conversion_ratio",
            "conversion": "feed_conversion_ratio",
        }

        for pattern, normalized in metric_mappings.items():
            if pattern in header_lower:
                return normalized

        cleaned = re.sub(r"[_-]+", "_", header_lower)
        cleaned = re.sub(r"[^\w_]", "", cleaned)
        return cleaned

    def _extract_pharmaceutical_data(
        self, sheet: Worksheet, sheet_name: str
    ) -> List[MetricData]:
        """Extraction données pharmaceutiques"""
        metrics = []
        category = "pharmaceutical"

        header_row, headers = self._find_headers_row(sheet)
        if not headers:
            return self._extract_generic_data(sheet, sheet_name)

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row = sheet[row_idx]
            if not row[0].value:
                break

            primary_key = str(row[0].value).strip()

            for col_idx, header in enumerate(headers):
                if col_idx >= len(row) or not row[col_idx].value:
                    continue

                if col_idx == 0:
                    metric = MetricData(
                        sheet_name=sheet_name,
                        category=category,
                        metric_key=f"pharma_ref_{primary_key}",
                        metric_name=f"Reference: {primary_key}",
                        value_text=str(row[col_idx].value),
                        metadata={
                            "format": "pharmaceutical",
                            "data_type": "pharmaceutical",
                            "is_reference_column": True,
                            "original_header": header,
                        },
                    )
                    metrics.append(metric)
                    continue

                metric_key = f"{primary_key}_{header.lower()}"
                metric_name = f"{header} for {primary_key}"

                value_raw = str(row[col_idx].value)
                value_numeric, unit = self._parse_numeric_value(value_raw)
                age_min, age_max = self._parse_age_range(header, value_raw)

                metric = MetricData(
                    sheet_name=sheet_name,
                    category=category,
                    metric_key=metric_key,
                    metric_name=metric_name,
                    value_text=value_raw,
                    value_numeric=value_numeric,
                    unit=unit,
                    age_min=age_min,
                    age_max=age_max,
                    metadata={
                        "format": "pharmaceutical",
                        "data_type": "pharmaceutical",
                        "drug_name": primary_key,
                        "original_header": header,
                    },
                )

                metrics.append(metric)

        return metrics

    def _extract_nutrition_data(
        self, sheet: Worksheet, sheet_name: str
    ) -> List[MetricData]:
        """Extraction données nutritionnelles"""
        metrics = []
        category = "nutrition"

        header_row, headers = self._find_headers_row(sheet)
        if not headers:
            return self._extract_generic_data(sheet, sheet_name)

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row = sheet[row_idx]
            if not row[0].value:
                break

            nutrient_name = str(row[0].value).strip()

            for col_idx, header in enumerate(headers):
                if col_idx >= len(row) or not row[col_idx].value:
                    continue

                if col_idx == 0:
                    metric = MetricData(
                        sheet_name=sheet_name,
                        category=category,
                        metric_key=f"nutrient_{nutrient_name}",
                        metric_name=f"Nutrient: {nutrient_name}",
                        value_text=str(row[col_idx].value),
                        metadata={
                            "format": "nutrition",
                            "data_type": "nutrition",
                            "is_nutrient_name": True,
                            "original_header": header,
                        },
                    )
                    metrics.append(metric)
                    continue

                metric_key = f"{nutrient_name}_{header.lower()}"
                metric_name = f"{header} of {nutrient_name}"

                value_raw = str(row[col_idx].value)
                value_numeric, unit = self._parse_numeric_value(value_raw)

                metric = MetricData(
                    sheet_name=sheet_name,
                    category=category,
                    metric_key=metric_key,
                    metric_name=metric_name,
                    value_text=value_raw,
                    value_numeric=value_numeric,
                    unit=unit,
                    metadata={
                        "format": "nutrition",
                        "data_type": "nutrition",
                        "nutrient": nutrient_name,
                        "original_header": header,
                    },
                )

                metrics.append(metric)

        return metrics

    def _extract_carcass_data(
        self, sheet: Worksheet, sheet_name: str
    ) -> List[MetricData]:
        """Extraction données carcasse"""
        metrics = []
        category = "carcass"

        header_row, headers = self._find_headers_row(sheet)
        if not headers:
            return self._extract_generic_data(sheet, sheet_name)

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row = sheet[row_idx]
            if not row[0].value:
                break

            primary_value = str(row[0].value).strip()
            age_min, age_max = self._parse_age_range(primary_value, primary_value)

            for col_idx, header in enumerate(headers):
                if col_idx >= len(row) or not row[col_idx].value:
                    continue

                if col_idx == 0:
                    if age_min is not None:
                        metric = MetricData(
                            sheet_name=sheet_name,
                            category=category,
                            metric_key=f"carcass_age_day_{age_min}",
                            metric_name=f"Carcass processing age: day {age_min}",
                            value_text=str(row[col_idx].value),
                            value_numeric=float(age_min),
                            unit="days",
                            age_min=age_min,
                            age_max=age_max,
                            metadata={
                                "format": "carcass",
                                "data_type": "carcass",
                                "is_age_reference": True,
                                "original_header": header,
                            },
                        )
                    else:
                        metric = MetricData(
                            sheet_name=sheet_name,
                            category=category,
                            metric_key=f"carcass_ref_{primary_value}",
                            metric_name=f"Carcass reference: {primary_value}",
                            value_text=str(row[col_idx].value),
                            metadata={
                                "format": "carcass",
                                "data_type": "carcass",
                                "is_reference_column": True,
                                "original_header": header,
                            },
                        )

                    metrics.append(metric)
                    continue

                metric_key = f"{primary_value}_{header.lower()}"
                metric_name = f"{header} at {primary_value}"

                value_raw = str(row[col_idx].value)
                value_numeric, unit = self._parse_numeric_value(value_raw)

                metric = MetricData(
                    sheet_name=sheet_name,
                    category=category,
                    metric_key=metric_key,
                    metric_name=metric_name,
                    value_text=value_raw,
                    value_numeric=value_numeric,
                    unit=unit,
                    age_min=age_min,
                    age_max=age_max,
                    metadata={
                        "format": "carcass",
                        "data_type": "carcass",
                        "original_header": header,
                    },
                )

                metrics.append(metric)

        return metrics

    def _extract_generic_data(
        self, sheet: Worksheet, sheet_name: str
    ) -> List[MetricData]:
        """Extraction générique"""
        potential_headers = []
        header_row_idx = None

        for row_idx in range(1, min(11, sheet.max_row + 1)):
            row = sheet[row_idx]
            row_values = [str(cell.value).strip() for cell in row[:10] if cell.value]

            if len(row_values) >= 3 and any(
                keyword in " ".join(row_values).lower()
                for keyword in [
                    "day",
                    "week",
                    "age",
                    "weight",
                    "gain",
                    "fcr",
                    "intake",
                    "conversion",
                ]
            ):
                potential_headers = row_values
                header_row_idx = row_idx
                logger.info(
                    f"En-têtes potentiels détectés ligne {row_idx}: {potential_headers}"
                )
                break

        if potential_headers and header_row_idx:
            logger.info("Structure tabulaire détectée, extraction dirigée")
            return self._extract_tabular_like_generic(
                sheet, sheet_name, header_row_idx, potential_headers
            )
        else:
            logger.info("Aucune structure tabulaire, extraction cellule par cellule")
            return self._extract_cell_by_cell_generic(sheet, sheet_name)

    def _extract_tabular_like_generic(
        self, sheet: Worksheet, sheet_name: str, header_row: int, headers: List[str]
    ) -> List[MetricData]:
        """Extraction tabulaire générique"""
        metrics = []
        category = self._categorize_sheet(sheet_name)

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row = sheet[row_idx]
            if not row[0].value:
                break

            first_value = str(row[0].value).strip()
            age_detected = None

            try:
                age_detected = int(float(first_value))
            except (ValueError, TypeError):
                age_match = re.search(r"(\d+)", first_value)
                if age_match:
                    age_detected = int(age_match.group(1))

            for col_idx, header in enumerate(headers):
                if col_idx >= len(row) or not row[col_idx].value:
                    continue

                cell = row[col_idx]

                if col_idx == 0:
                    if age_detected is not None:
                        metric = MetricData(
                            sheet_name=sheet_name,
                            category=category,
                            metric_key=f"generic_age_day_{age_detected}",
                            metric_name=f"Age at day {age_detected}",
                            value_text=str(cell.value),
                            value_numeric=float(age_detected),
                            unit="days",
                            age_min=age_detected,
                            age_max=age_detected,
                            metadata={
                                "format": "generic_tabular",
                                "is_age_column": True,
                                "row": row_idx,
                                "col": col_idx,
                                "original_header": header,
                            },
                        )
                    else:
                        metric = MetricData(
                            sheet_name=sheet_name,
                            category=category,
                            metric_key=f"generic_ref_{first_value}",
                            metric_name=f"Reference: {first_value}",
                            value_text=str(cell.value),
                            metadata={
                                "format": "generic_tabular",
                                "is_reference_column": True,
                                "row": row_idx,
                                "col": col_idx,
                                "original_header": header,
                            },
                        )

                    metrics.append(metric)
                    continue

                if age_detected is not None:
                    metric_key = f"generic_day_{age_detected}_{header.lower()}"
                    metric_name = f"{header} at day {age_detected}"
                    age_min, age_max = age_detected, age_detected
                else:
                    metric_key = f"generic_{first_value}_{header.lower()}"
                    metric_name = f"{header} for {first_value}"
                    age_min, age_max = None, None

                value_numeric, unit = self._parse_numeric_value(str(cell.value))

                metric = MetricData(
                    sheet_name=sheet_name,
                    category=category,
                    metric_key=metric_key,
                    metric_name=metric_name,
                    value_text=str(cell.value),
                    value_numeric=value_numeric,
                    unit=unit,
                    age_min=age_min,
                    age_max=age_max,
                    metadata={
                        "format": "generic_tabular",
                        "row": row_idx,
                        "col": col_idx,
                        "original_header": header,
                    },
                )

                metrics.append(metric)

        return metrics

    def _extract_cell_by_cell_generic(
        self, sheet: Worksheet, sheet_name: str
    ) -> List[MetricData]:
        """Extraction cellule par cellule"""
        metrics = []
        category = self._categorize_sheet(sheet_name)

        for row_idx, row in enumerate(sheet.iter_rows(max_row=100), 1):
            for col_idx, cell in enumerate(row):
                if cell.value and isinstance(cell.value, (int, float)):
                    metric = MetricData(
                        sheet_name=sheet_name,
                        category=category,
                        metric_key=f"cell_R{row_idx}C{col_idx+1}",
                        metric_name=f"Value at R{row_idx}C{col_idx+1}",
                        value_numeric=float(cell.value),
                        value_text=str(cell.value),
                        metadata={
                            "format": "generic_cell",
                            "row": row_idx,
                            "col": col_idx + 1,
                        },
                    )
                    metrics.append(metric)

        return metrics

    def _find_headers_row(self, sheet: Worksheet) -> Tuple[int, List[str]]:
        """Trouve la ligne d'en-têtes dans une feuille"""
        for row_idx in range(1, min(30, sheet.max_row + 1)):
            row = sheet[row_idx]
            headers = []

            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    header = str(cell.value).strip()
                    if header:
                        headers.append(header)
                else:
                    break

            if len(headers) >= 3 and any(
                keyword in " ".join(headers).lower()
                for keyword in [
                    "day",
                    "week",
                    "age",
                    "weight",
                    "gain",
                    "fcr",
                    "drug",
                    "nutrient",
                    "yield",
                ]
            ):
                return row_idx, headers

        return 0, []

    def _has_metadata_format(self, sheet: Worksheet) -> bool:
        """Vérifie si la feuille utilise le format metadata/value"""
        try:
            return (
                sheet["A1"].value
                and sheet["B1"].value
                and str(sheet["A1"].value).lower() == "metadata"
                and str(sheet["B1"].value).lower() == "value"
            )
        except Exception:
            return False

    def _categorize_sheet(self, sheet_name: str) -> str:
        """Catégorise une feuille selon son nom"""
        sheet_lower = sheet_name.lower()

        if self.data_type in self.category_mapping:
            return self.data_type

        for category, keywords in self.category_mapping.items():
            if any(keyword in sheet_lower for keyword in keywords):
                return category

        return "other"

    def _parse_numeric_value(
        self, value_str: str
    ) -> Tuple[Optional[float], Optional[str]]:
        """Extrait valeur numérique et unité"""
        if not value_str:
            return None, None

        patterns = [
            r"(\d+\.?\d*)\s*(%|kg|g|mg|ml|l|cm|mm|°C|°F|hrs?|days?|weeks?|months?|ppm|IU)",
            r"(\d+\.?\d*)",
        ]

        for pattern in patterns:
            match = re.search(pattern, str(value_str))
            if match:
                try:
                    value = float(match.group(1))
                    unit = match.group(2) if len(match.groups()) > 1 else None
                    return value, unit
                except ValueError:
                    continue

        return None, None

    def _parse_age_range(
        self, key: str, value: str
    ) -> Tuple[Optional[int], Optional[int]]:
        """Extrait plage d'âge"""
        text = f"{key} {value}".lower()

        patterns = [
            r"(\d+)-(\d+)\s*(?:days?|jours?)",
            r"(\d+)-(\d+)\s*(?:weeks?|semaines?)",
            r"(?:day|jour)\s*(\d+)",
            r"(\d+)\s*(?:days?|jours?)",
            r"(?:week|semaine)\s*(\d+)",
            r"(\d+)\s*(?:weeks?|semaines?)",
        ]

        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                if len(match.groups()) == 2:
                    return int(match.group(1)), int(match.group(2))
                else:
                    age = int(match.group(1))
                    if "week" in pattern or "semaine" in pattern:
                        age *= 7
                    return age, age

        return None, None

    def _clean_metric_name(self, metric_key: str) -> str:
        """Nettoie le nom de métrique"""
        cleaned = re.sub(r"[_-]+", " ", metric_key)
        cleaned = re.sub(r"\s+", " ", cleaned)
        return cleaned.title().strip()

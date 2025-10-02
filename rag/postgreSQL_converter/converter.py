#!/usr/bin/env python3
"""
Convertisseur Excel Intelligent vers PostgreSQL
Point d'entrée principal avec orchestration des modules
"""

import asyncio
import sys
import logging
import hashlib
from pathlib import Path
from typing import Dict, Any

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print("ERREUR: openpyxl requis. Installez avec: pip install openpyxl")
    sys.exit(1)

from config import DATABASE_CONFIG, validate_database_config, IntentsConfigLoader
from models import TaxonomyInfo
from format_detector import EnhancedFormatDetector
from data_extractor import IntelligentDataExtractor
from database import PostgreSQLManager

# Configuration logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


class IntelligentExcelConverter:
    """Convertisseur Excel intelligent avec support intents.json"""

    def __init__(self, db_config: Dict[str, Any], intents_config_path: str = None):
        self.intents_config = IntentsConfigLoader(intents_config_path)
        self.format_detector = EnhancedFormatDetector(self.intents_config)
        self.db_manager = PostgreSQLManager(db_config)

    async def initialize(self):
        await self.db_manager.initialize()
        logger.info(
            "Convertisseur intelligent initialisé avec support intents.json v1.2"
        )

    async def convert_file(self, file_path: str, force_reprocess: bool = False) -> bool:
        """Convertit un fichier avec détection intelligente et support feuilles multiples"""

        try:
            logger.info(f"Conversion intelligente: {file_path}")
            filename = Path(file_path).name

            # Calcul hash
            with open(file_path, "rb") as f:
                current_file_hash = hashlib.md5(f.read()).hexdigest()

            # Chargement et analyse
            workbook = load_workbook(file_path, read_only=True, data_only=True)

            # Détection format et type global avec intents.json
            format_type, data_type, global_taxonomy = (
                self.format_detector.detect_format_and_type(workbook, filename)
            )

            # Vérifier si le fichier contient des feuilles avec métadonnées individuelles
            metadata_sheets = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                if self._has_metadata_format(sheet):
                    metadata_sheets.append(sheet_name)

            total_documents_created = 0
            total_metrics_created = 0

            if metadata_sheets:
                # Mode multi-feuilles : créer un document par feuille avec métadonnées
                logger.info(
                    f"Mode multi-feuilles détecté: {len(metadata_sheets)} feuilles avec métadonnées"
                )

                for sheet_name in metadata_sheets:
                    sheet = workbook[sheet_name]

                    # Extraire taxonomie spécifique à cette feuille
                    sheet_taxonomy = self._extract_sheet_specific_taxonomy(
                        sheet, sheet_name, global_taxonomy
                    )

                    # Extraire métriques pour cette feuille uniquement
                    extractor = IntelligentDataExtractor(
                        format_type, data_type, self.intents_config
                    )
                    sheet_metrics = extractor._extract_metadata_sheet(sheet, sheet_name)

                    if sheet_metrics:
                        # Créer un nom de fichier unique pour cette feuille
                        sheet_filename = f"{filename}#{sheet_name}"
                        sheet_file_hash = hashlib.md5(
                            f"{current_file_hash}#{sheet_name}".encode()
                        ).hexdigest()

                        # Transférer métadonnées si disponibles
                        if hasattr(extractor, "_current_table_metadata"):
                            self.db_manager._current_table_metadata = (
                                extractor._current_table_metadata
                            )

                        # Insertion en base pour cette feuille
                        document_id = await self.db_manager.insert_document_data(
                            sheet_taxonomy,
                            sheet_metrics,
                            sheet_filename,
                            sheet_file_hash,
                        )

                        total_documents_created += 1
                        total_metrics_created += len(sheet_metrics)

                        logger.info(
                            f"Document créé pour feuille '{sheet_name}' - ID: {document_id}"
                        )
                        logger.info(f"  - Sexe: {sheet_taxonomy.sex}")
                        logger.info(f"  - Métriques: {len(sheet_metrics)}")
                    else:
                        logger.warning(
                            f"Aucune métrique extraite de la feuille '{sheet_name}'"
                        )

                if total_documents_created > 0:
                    logger.info("Conversion multi-feuilles réussie:")
                    logger.info(f"  - Documents créés: {total_documents_created}")
                    logger.info(f"  - Métriques totales: {total_metrics_created}")
                    logger.info(f"  - Format: {format_type}, Type: {data_type}")
                    workbook.close()
                    return True
                else:
                    logger.error("Aucun document créé en mode multi-feuilles")
                    workbook.close()
                    return False

            else:
                # Mode fichier unique : traitement classique
                logger.info("Mode fichier unique: traitement global")

                extractor = IntelligentDataExtractor(
                    format_type, data_type, self.intents_config
                )
                metrics = extractor.extract_metrics(workbook)

                if not metrics:
                    logger.warning(f"Aucune métrique extraite de {filename}")
                    workbook.close()
                    return False

                # Transférer les métadonnées de table vers le gestionnaire de base de données
                if hasattr(extractor, "_current_table_metadata"):
                    self.db_manager._current_table_metadata = (
                        extractor._current_table_metadata
                    )

                # Insertion en base
                document_id = await self.db_manager.insert_document_data(
                    global_taxonomy, metrics, filename, current_file_hash
                )

                logger.info(f"Conversion réussie - Document ID: {document_id}")
                logger.info(
                    f"Format: {format_type}, Type: {data_type}, Métriques: {len(metrics)}"
                )

                # Log des métriques importantes pour debug
                important_metrics = [
                    m
                    for m in metrics
                    if "weight" in m.metric_name.lower()
                    or "fcr" in m.metric_name.lower()
                ]
                if important_metrics:
                    logger.info("Métriques importantes extraites:")
                    for metric in important_metrics[:5]:  # Limiter à 5 exemples
                        logger.info(
                            f"  - {metric.metric_name}: {metric.value_numeric} {metric.unit or ''}"
                        )

                workbook.close()
                return True

        except Exception as e:
            logger.error(f"Erreur conversion {file_path}: {e}")
            raise

    async def close(self):
        await self.db_manager.close()

    def _has_metadata_format(self, sheet: Worksheet) -> bool:
        """Vérifie si la feuille a un format metadata/value"""
        try:
            cell_a1 = sheet["A1"].value
            cell_b1 = sheet["B1"].value
            return (
                cell_a1
                and cell_b1
                and str(cell_a1).lower() == "metadata"
                and str(cell_b1).lower() == "value"
            )
        except Exception:
            return False

    def _extract_sheet_specific_taxonomy(
        self, sheet: Worksheet, sheet_name: str, global_taxonomy: TaxonomyInfo
    ) -> TaxonomyInfo:
        """Extrait la taxonomie spécifique à une feuille depuis ses métadonnées"""

        # Extraire les métadonnées de cette feuille
        metadata_pairs = {}
        try:
            for row in sheet.iter_rows(min_row=2, max_row=50):
                if row[0].value and row[1].value:
                    key = str(row[0].value).lower().strip()
                    value = str(row[1].value).strip()
                    metadata_pairs[key] = value
        except Exception:
            pass

        # Extraire le sexe spécifique de cette feuille
        sheet_sex = None
        if "sex" in metadata_pairs:
            sheet_sex = self.format_detector._normalize_sex_value(metadata_pairs["sex"])
            logger.info(
                f"Sexe détecté pour feuille '{sheet_name}': {metadata_pairs['sex']} -> {sheet_sex}"
            )

        # Extraire autres informations spécifiques si disponibles
        housing_system = metadata_pairs.get(
            "housing_system", global_taxonomy.housing_system
        )
        feather_color = metadata_pairs.get(
            "feather_type", global_taxonomy.feather_color
        )

        # Déterminer le nom de strain plus spécifique si possible
        strain_name = global_taxonomy.strain
        if "strain" in metadata_pairs:
            strain_name = metadata_pairs["strain"]
        elif "name" in metadata_pairs:
            # Utiliser le nom de la feuille comme strain plus spécifique
            strain_name = f"{global_taxonomy.strain} - {metadata_pairs['name']}"

        # Créer taxonomie spécifique à cette feuille
        sheet_taxonomy = TaxonomyInfo(
            company=global_taxonomy.company,
            breed=global_taxonomy.breed,
            strain=strain_name,
            species=global_taxonomy.species,
            housing_system=housing_system,
            feather_color=feather_color,
            sex=sheet_sex,  # Sexe spécifique à cette feuille
            data_type=global_taxonomy.data_type,
        )

        logger.debug(
            f"Taxonomie pour feuille '{sheet_name}': {sheet_taxonomy.sex}, {sheet_taxonomy.strain}"
        )

        return sheet_taxonomy


async def main():
    """Point d'entrée principal avec support intents.json"""

    # Validation configuration
    try:
        validate_database_config()
    except ValueError as e:
        logger.error(str(e))
        sys.exit(1)

    converter = IntelligentExcelConverter(DATABASE_CONFIG)

    try:
        await converter.initialize()

        if len(sys.argv) >= 2:
            file_path = sys.argv[1]
            success = await converter.convert_file(file_path)
            print(f"Conversion {'réussie' if success else 'échouée'}")
        else:
            print("Usage: python converter.py <fichier.xlsx>")
            print(
                "Le fichier intents.json sera automatiquement détecté s'il est présent"
            )

    finally:
        await converter.close()


if __name__ == "__main__":
    asyncio.run(main())

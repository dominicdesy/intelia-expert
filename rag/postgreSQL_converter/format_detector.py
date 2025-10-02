#!/usr/bin/env python3
"""
Détecteur de formats Excel avec support intents.json
"""

import re
import logging
from pathlib import Path
from typing import List, Tuple, Optional, Dict

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from models import TaxonomyInfo
from config import IntentsConfigLoader

logger = logging.getLogger(__name__)


class EnhancedFormatDetector:
    """Détecteur intelligent multi-format avec support intents.json"""

    def __init__(self, intents_config: Optional[IntentsConfigLoader] = None):
        self.intents_config = intents_config or IntentsConfigLoader()
        self.line_aliases = self.intents_config.get_line_aliases()
        self.metric_types = self.intents_config.get_metric_types()

        # Patterns étendus avec support aliases
        self.format_patterns = {
            "ross_performance": {
                "file_patterns": ["ross", "aviagen"]
                + self._get_patterns_for_line("ross 308"),
                "sheet_indicators": ["male_metric", "female_metric", "as_hatched"],
                "header_patterns": ["Day", "Weight_g", "FCR", "Daily_Gain"],
                "metadata_pattern": "metadata/value",
                "data_type": "performance",
            },
            "cobb_performance": {
                "file_patterns": ["cobb"] + self._get_patterns_for_line("cobb 500"),
                "sheet_indicators": ["male", "female", "mixed"],
                "header_patterns": ["week", "body_weight", "feed_conversion"],
                "metadata_pattern": "week/metric",
                "data_type": "performance",
            },
            "hyline_performance": {
                "file_patterns": ["hyline", "hy-line"]
                + self._get_patterns_for_line("hy-line brown"),
                "sheet_indicators": ["brown", "white"],
                "header_patterns": ["age", "body_weight", "egg_production"],
                "metadata_pattern": "metadata/value",
                "data_type": "performance",
            },
            "pharmaceutical": {
                "file_patterns": [
                    "vaccine",
                    "drug",
                    "medication",
                    "pharma",
                    "treatment",
                    "protocol",
                ],
                "sheet_indicators": ["dosage", "protocol", "schedule", "vaccination"],
                "header_patterns": ["drug_name", "dosage", "age_application", "route"],
                "metadata_pattern": "drug/dosage",
                "data_type": "pharmaceutical",
            },
            "nutrition": {
                "file_patterns": [
                    "feed",
                    "nutrition",
                    "diet",
                    "amino",
                    "vitamin",
                    "mineral",
                ],
                "sheet_indicators": ["starter", "grower", "finisher", "layer"],
                "header_patterns": ["nutrient", "content", "percentage", "requirement"],
                "metadata_pattern": "nutrient/value",
                "data_type": "nutrition",
            },
            "carcass": {
                "file_patterns": ["carcass", "yield", "processing", "slaughter"],
                "sheet_indicators": ["carcass_male", "carcass_female", "parts"],
                "header_patterns": [
                    "live_weight",
                    "carcass_weight",
                    "yield_percentage",
                ],
                "metadata_pattern": "weight/yield",
                "data_type": "carcass",
            },
        }

    def _get_patterns_for_line(self, line_name: str) -> List[str]:
        """Récupère tous les alias pour une lignée génétique"""
        aliases = self.line_aliases.get(line_name, [])
        cleaned_aliases = []
        for alias in aliases:
            clean = re.sub(r"[^\w\s]", "", alias.lower())
            cleaned_aliases.append(clean)
        return cleaned_aliases

    def detect_format_and_type(
        self, workbook: openpyxl.Workbook, filename: str
    ) -> Tuple[str, str, TaxonomyInfo]:
        """Détection améliorée avec reconnaissance des lignées via intents.json"""
        logger.info(f"Détection intelligente pour: {filename}")

        filename_lower = filename.lower()
        detected_format = "generic"
        data_type = "other"

        format_scores = {}

        for format_name, patterns in self.format_patterns.items():
            score = 0

            for pattern in patterns["file_patterns"]:
                if pattern in filename_lower:
                    score += 3
                    logger.debug(
                        f"Pattern '{pattern}' trouvé dans filename pour {format_name}"
                    )

            for sheet_name in workbook.sheetnames:
                sheet_lower = sheet_name.lower()
                for indicator in patterns["sheet_indicators"]:
                    if indicator in sheet_lower:
                        score += 2
                        logger.debug(
                            f"Indicator '{indicator}' trouvé dans sheet '{sheet_name}' pour {format_name}"
                        )

            for sheet_name in workbook.sheetnames[:3]:
                try:
                    sheet = workbook[sheet_name]
                    headers_found = self._extract_headers_from_sheet(sheet)
                    for pattern in patterns["header_patterns"]:
                        if any(
                            pattern.lower() in header.lower()
                            for header in headers_found
                        ):
                            score += 1
                            logger.debug(
                                f"Header pattern '{pattern}' trouvé pour {format_name}"
                            )
                except Exception:
                    continue

            format_scores[format_name] = score
            logger.debug(f"Score pour {format_name}: {score}")

        if format_scores:
            best_format = max(format_scores, key=format_scores.get)
            if format_scores[best_format] > 0:
                detected_format = best_format
                data_type = self.format_patterns[best_format]["data_type"]

        taxonomy = self._extract_enhanced_taxonomy(
            workbook, detected_format, filename, data_type
        )

        logger.info(
            f"Format détecté: {detected_format} (score: {format_scores.get(detected_format, 0)})"
        )
        logger.info(f"Type de données: {data_type}")
        logger.info(
            f"Taxonomie: {taxonomy.company} - {taxonomy.breed} - {taxonomy.strain}"
        )

        return detected_format, data_type, taxonomy

    def _extract_headers_from_sheet(self, sheet: Worksheet) -> List[str]:
        """Extrait les en-têtes potentiels d'une feuille"""
        headers = []

        for row_idx in range(1, min(21, sheet.max_row + 1)):
            row_headers = []
            for col_idx in range(1, min(11, sheet.max_column + 1)):
                cell = sheet.cell(row_idx, col_idx)
                if cell.value and isinstance(cell.value, str):
                    row_headers.append(str(cell.value).strip())

            if len(row_headers) >= 3:
                headers.extend(row_headers)

        return headers

    def _extract_enhanced_taxonomy(
        self,
        workbook: openpyxl.Workbook,
        format_type: str,
        filename: str,
        data_type: str,
    ) -> TaxonomyInfo:
        """Extraction taxonomique améliorée avec reconnaissance lignées via intents.json"""

        if format_type.startswith("ross"):
            return self._extract_ross_taxonomy(workbook, filename, data_type)
        elif format_type.startswith("cobb"):
            return self._extract_cobb_taxonomy(workbook, filename, data_type)
        elif format_type.startswith("hyline"):
            return self._extract_hyline_taxonomy(workbook, filename, data_type)
        elif format_type == "pharmaceutical":
            return self._extract_pharmaceutical_taxonomy(workbook, filename)
        elif format_type == "nutrition":
            return self._extract_nutrition_taxonomy(workbook, filename)
        else:
            return self._extract_intelligent_taxonomy(workbook, filename, data_type)

    def _extract_intelligent_taxonomy(
        self, workbook: openpyxl.Workbook, filename: str, data_type: str
    ) -> TaxonomyInfo:
        """Extraction intelligente basée sur intents.json"""

        filename_lower = filename.lower()
        detected_line = self._detect_genetic_line(filename_lower)
        company, breed, species = self._map_line_to_taxonomy(detected_line)
        detected_sex = self._detect_sex(filename_lower)

        return TaxonomyInfo(
            company=company,
            breed=breed,
            strain=(
                detected_line
                if detected_line
                else self._extract_strain_from_filename(filename)
            ),
            species=species,
            sex=detected_sex,
            data_type=data_type,
        )

    def _detect_genetic_line(self, text: str) -> Optional[str]:
        """Détecte la lignée génétique via les alias d'intents.json"""

        for line_name, aliases in self.line_aliases.items():
            if line_name.lower() in text:
                return line_name

            for alias in aliases:
                clean_alias = re.sub(r"[^\w\s]", "", alias.lower())
                if clean_alias in text:
                    logger.debug(f"Lignée détectée: {line_name} via alias '{alias}'")
                    return line_name

        return None

    def _detect_sex(self, text: str) -> Optional[str]:
        """Détecte le sexe via les alias d'intents.json"""

        sex_aliases = self.intents_config.get_sex_aliases()

        for sex_name, aliases in sex_aliases.items():
            if sex_name.lower() in text:
                return sex_name

            for alias in aliases:
                if alias.lower() in text:
                    logger.debug(f"Sexe détecté: {sex_name} via alias '{alias}'")
                    return sex_name

        return None

    def _extract_sex_from_sheets(self, workbook: openpyxl.Workbook) -> Optional[str]:
        """Extrait le sexe depuis les métadonnées des feuilles"""

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            if self._has_metadata_format(sheet):
                metadata_pairs = self._extract_metadata_pairs(sheet)

                for key, value in metadata_pairs.items():
                    if key in ["sex", "sexe"]:
                        normalized_sex = self._normalize_sex_value(value)
                        if normalized_sex:
                            logger.info(
                                f"Sexe trouvé dans feuille {sheet_name}: {value} -> {normalized_sex}"
                            )
                            return normalized_sex

        return None

    def _normalize_sex_value(self, sex_value: str) -> Optional[str]:
        """Normalise une valeur de sexe via intents.json"""

        sex_aliases = self.intents_config.get_sex_aliases()
        value_lower = sex_value.lower().strip()

        for sex_name, aliases in sex_aliases.items():
            if value_lower == sex_name.lower():
                return sex_name

            for alias in aliases:
                if value_lower == alias.lower():
                    logger.debug(
                        f"Sexe normalisé: {sex_value} -> {sex_name} via alias '{alias}'"
                    )
                    return sex_name

        fallback_patterns = {
            "male": ["m", "male", "mâle", "masculin", "coq", "rooster"],
            "female": ["f", "female", "femelle", "féminin", "poule", "hen", "laying"],
            "mixed": ["mix", "mixed", "mixte", "as hatched", "straight run", "unsexed"],
        }

        for sex_type, patterns in fallback_patterns.items():
            if value_lower in patterns:
                logger.debug(f"Sexe normalisé via fallback: {sex_value} -> {sex_type}")
                return sex_type

        logger.warning(f"Valeur de sexe non reconnue: '{sex_value}'")
        return None

    def _map_line_to_taxonomy(self, line_name: Optional[str]) -> Tuple[str, str, str]:
        """Mappe une lignée génétique vers company/breed/species"""

        if not line_name:
            return "Unknown", "Unknown", "broiler"

        line_lower = line_name.lower()

        if "ross" in line_lower:
            return "Aviagen", "Ross", "broiler"
        elif "cobb" in line_lower:
            return "Cobb-Vantress", "Cobb", "broiler"
        elif "hy-line" in line_lower or "hyline" in line_lower:
            return "EW Group", "Hy-Line", "layer"
        elif any(
            term in line_lower for term in ["hubbard", "ranger", "sasso", "freedom"]
        ):
            return "Hubbard", "Alternative", "broiler"
        elif any(
            term in line_lower
            for term in ["isa", "lohmann", "dekalb", "bovans", "shaver"]
        ):
            return "Layer Genetics", "Commercial Layer", "layer"

        species = (
            "layer"
            if any(term in line_lower for term in ["layer", "laying", "brown", "white"])
            else "broiler"
        )

        return "Unknown", "Mixed", species

    def _extract_ross_taxonomy(
        self, workbook: openpyxl.Workbook, filename: str, data_type: str
    ) -> TaxonomyInfo:
        """Extraction spécifique Ross/Aviagen"""

        detected_sex = self._extract_sex_from_sheets(workbook)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if self._has_metadata_format(sheet):
                taxonomy_data = self._extract_metadata_pairs(sheet)

                return TaxonomyInfo(
                    company=taxonomy_data.get("brand", "Aviagen"),
                    breed=taxonomy_data.get("breed", "Ross"),
                    strain=taxonomy_data.get(
                        "strain", self._extract_strain_from_filename(filename)
                    ),
                    species="broiler",
                    housing_system=taxonomy_data.get("housing_system"),
                    feather_color=taxonomy_data.get("feather_type"),
                    sex=detected_sex,
                    data_type=data_type,
                )

        return TaxonomyInfo(
            company="Aviagen",
            breed="Ross",
            strain=self._extract_strain_from_filename(filename),
            species="broiler",
            sex=detected_sex,
            data_type=data_type,
        )

    def _extract_cobb_taxonomy(
        self, workbook: openpyxl.Workbook, filename: str, data_type: str
    ) -> TaxonomyInfo:
        """Extraction spécifique Cobb"""
        return TaxonomyInfo(
            company="Cobb-Vantress",
            breed="Cobb",
            strain=self._extract_strain_from_filename(filename),
            species="broiler",
            data_type=data_type,
        )

    def _extract_hyline_taxonomy(
        self, workbook: openpyxl.Workbook, filename: str, data_type: str
    ) -> TaxonomyInfo:
        """Extraction spécifique Hyline"""

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if self._has_metadata_format(sheet):
                taxonomy_data = self._extract_metadata_pairs(sheet)

                return TaxonomyInfo(
                    company=taxonomy_data.get("brand", "EW Group"),
                    breed=taxonomy_data.get("breed", "Hy-Line"),
                    strain=taxonomy_data.get(
                        "strain", self._extract_strain_from_filename(filename)
                    ),
                    species="layer",
                    housing_system=taxonomy_data.get("housing_system"),
                    feather_color=taxonomy_data.get("feather_color"),
                    sex=taxonomy_data.get("sex"),
                    data_type=data_type,
                )

        return TaxonomyInfo(
            company="EW Group",
            breed="Hy-Line",
            strain=self._extract_strain_from_filename(filename),
            species="layer",
            data_type=data_type,
        )

    def _extract_pharmaceutical_taxonomy(
        self, workbook: openpyxl.Workbook, filename: str
    ) -> TaxonomyInfo:
        """Extraction pour données pharmaceutiques"""

        species = "broiler"
        filename_lower = filename.lower()

        if any(term in filename_lower for term in ["layer", "laying", "hen", "egg"]):
            species = "layer"

        return TaxonomyInfo(
            company="Pharmaceutical",
            breed="Multi-species",
            strain="Universal",
            species=species,
            data_type="pharmaceutical",
        )

    def _extract_nutrition_taxonomy(
        self, workbook: openpyxl.Workbook, filename: str
    ) -> TaxonomyInfo:
        """Extraction pour données nutritionnelles"""

        species = "broiler"
        filename_lower = filename.lower()

        if any(term in filename_lower for term in ["layer", "laying", "hen"]):
            species = "layer"

        return TaxonomyInfo(
            company="Nutrition",
            breed="Feed Requirements",
            strain="Universal",
            species=species,
            data_type="nutrition",
        )

    def _has_metadata_format(self, sheet: Worksheet) -> bool:
        """Vérifie si la feuille a un format metadata/value"""
        try:
            cell_a1 = sheet["A1"].value
            cell_b1 = sheet["B1"].value
            return (
                cell_a1
                and cell_b1
                and str(cell_a1).lower() == "metadata"
                and str(cell_b1).lower() == "value"
            )
        except Exception:
            return False

    def _extract_metadata_pairs(self, sheet: Worksheet) -> Dict[str, str]:
        """Extrait les paires métadonnées/valeurs"""
        metadata = {}

        try:
            for row in sheet.iter_rows(min_row=2, max_row=30):
                if row[0].value and row[1].value:
                    key = str(row[0].value).lower().strip()
                    value = str(row[1].value).strip()
                    metadata[key] = value
        except Exception:
            pass

        return metadata

    def _extract_strain_from_filename(self, filename: str) -> str:
        """Extrait la souche du nom de fichier avec support intents.json"""

        detected_line = self._detect_genetic_line(filename.lower())
        if detected_line:
            return detected_line

        patterns = [
            r"(ross\s*\d+)",
            r"(cobb\s*\d+)",
            r"(\d+\s*ff?)",
            r"(brown\s+\w+)",
            r"(white\s+\w+)",
            r"(ap\s*\d+)",
            r"(\w+\s+alternative?)",
        ]

        filename_lower = filename.lower()

        for pattern in patterns:
            match = re.search(pattern, filename_lower)
            if match:
                return match.group(1).title()

        base_name = Path(filename).stem
        return re.sub(r"[^\w\s]", " ", base_name).title()
